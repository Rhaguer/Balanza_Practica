import json
import sqlite3
import threading
from contextlib import closing
from datetime import datetime, time, timedelta
from decimal import Decimal
from io import BytesIO, StringIO
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch
from zipfile import ZipFile
from openpyxl import load_workbook

from django.contrib.auth.models import Group, User
from django.core import mail
from django.core.cache import cache
from django.core.management import call_command
from django.db import OperationalError, close_old_connections, transaction
from django.db.models import Sum
from django.test import (
    Client,
    RequestFactory,
    SimpleTestCase,
    TestCase,
    TransactionTestCase,
    override_settings,
)
from django.urls import reverse
from django.utils import timezone

from .models import (
    Actividad,
    ClaseHorario,
    AuditLog,
    ESTADO_RESIDUO_ANULADO,
    ESTADO_RESIDUO_CONFIRMADO,
    ESTADO_RESIDUO_PENDIENTE,
    CategoriaResiduos,
    Destino,
    HistorialRetiro,
    OperationalLock,
    Residuo,
    TipoResiduos,
    Usuarios,
    WeightReading,
)
from .forms import ClaseHorarioForm, UsuariosForm
from .residue_types import TIPO_INORGANICO, TIPO_ORGANICO
from .roles import ROL_ADMINISTRADOR, ROL_ESTUDIANTE, ROL_PROFESOR
from .services import balanza_service
from .services.backup_service import (
    create_rolling_backup,
    create_verified_backup,
    is_sqlite_memory_database,
    restore_verified_backup,
    verify_sqlite_backup,
)
from .services.dashboard import (
    LIMITE_UNIDADES_GLOBAL,
    crear_dashboard_context,
    obtener_proximo_contenedor_disponible,
)
from .views import (
    _bloquear_capacidad,
    _respaldar_clases_por_finalizar,
    _validar_capacidad_unidades,
)


class FlujoResiduoTests(TestCase):
    def setUp(self):
        grupo, _ = Group.objects.get_or_create(name="Estudiante")
        self.user = User.objects.create_user(
            username="estudiante@test.cl",
            email="estudiante@test.cl",
            password="ClaveSegura123!",
        )
        self.user.groups.add(grupo)
        self.client.force_login(self.user)

        self.categoria = CategoriaResiduos.objects.create(
            nombre="Orgánico",
            descripcion="Residuos compostables",
            tipo_operacional=TIPO_ORGANICO,
        )
        self.subtipo = TipoResiduos.objects.create(
            categoria=self.categoria,
            nombre_residuo="Cáscaras",
            descripcion="Restos de fruta",
        )
        ahora = timezone.localtime()
        self.taller = ClaseHorario.objects.create(
            profesor="Profesor Demo",
            asignatura="Taller Compost",
            seccion="A1",
            horario="00:00 - 23:59",
            fecha=ahora.date(),
            dia_semana=ahora.weekday(),
            hora_inicio=time(0, 0),
            hora_fin=time(23, 59, 59),
        )

    def _crear_residuo_organico(self, contenedor, peso, numero_clase):
        return Residuo.objects.create(
            seccion="A1",
            profesor="Profesor Demo",
            asignatura="Taller Compost",
            horario="10:00 - 11:00",
            numero_clase=numero_clase,
            tipo=self.categoria,
            subtipo=self.subtipo,
            peso=Decimal(peso),
            contenedor_id=contenedor,
            estado=ESTADO_RESIDUO_CONFIRMADO,
            confirmado_at=timezone.now(),
            created_by=self.user,
        )

    def test_registro_inicia_pendiente_y_confirma_con_peso(self):
        response = self.client.post(reverse("guardar"), {
            "taller_id": self.taller.id,
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["success"])

        residuo = Residuo.objects.get(pk=payload["id"])
        self.assertEqual(residuo.estado, ESTADO_RESIDUO_PENDIENTE)
        self.assertEqual(residuo.numero_clase, 1)

        response = self.client.post(reverse("actualizar_residuo"), {
            "residuo_id": residuo.id,
            "tipo": self.categoria.id_categoria,
            "peso": "1.234",
        })

        self.assertEqual(response.status_code, 200)
        residuo.refresh_from_db()
        self.assertEqual(residuo.estado, ESTADO_RESIDUO_CONFIRMADO)
        self.assertEqual(residuo.peso, Decimal("1.234"))
        self.assertEqual(residuo.contenedor_id, "Compostera 1")
        self.assertIsNotNone(residuo.confirmado_at)
        self.assertEqual(
            response.json()["registro_contexto"]["asignatura"],
            "Taller Compost",
        )
        self.assertEqual(residuo.taller, self.taller)

    def test_pantalla_de_residuos_muestra_organicos_e_inorganicos_sin_subtipos(self):
        response = self.client.get(reverse("residuos"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Orgánicos")
        self.assertContains(response, "Inorgánicos")
        self.assertContains(response, 'data-tipo-operacional="organico"')
        self.assertContains(response, 'data-tipo-operacional="inorganico"')
        self.assertNotContains(response, "Subtipo")
        self.assertNotContains(response, "Líquidos")

    def test_dashboard_conserva_grafico_de_liquidos_deshabilitado(self):
        categoria_liquida_legacy = CategoriaResiduos.objects.create(
            nombre="Líquidos históricos",
            descripcion="Solo para comprobar compatibilidad del gráfico.",
            tipo_operacional="liquido",
        )
        subtipo_liquido_legacy = TipoResiduos.objects.create(
            categoria=categoria_liquida_legacy,
            nombre_residuo="Aceite histórico",
            descripcion="Registro heredado.",
        )
        Residuo.objects.create(
            seccion="A1",
            profesor="Profesor Demo",
            asignatura="Taller Compost",
            horario="10:00 - 11:00",
            numero_clase=2,
            tipo=categoria_liquida_legacy,
            subtipo=subtipo_liquido_legacy,
            peso=Decimal("40.000"),
            contenedor_id="Tanque Líquidos",
            estado=ESTADO_RESIDUO_CONFIRMADO,
            confirmado_at=timezone.now(),
            created_by=self.user,
        )

        context = crear_dashboard_context(self.user)
        response = self.client.get(reverse("dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertFalse(context["datos_liquidos"]["habilitado"])
        self.assertEqual(context["datos_liquidos"]["capacidad"], 200.0)
        self.assertEqual(context["datos_liquidos"]["total_litros"], 40.0)
        self.assertEqual(context["datos_liquidos"]["porcentaje"], 20.0)
        self.assertEqual(
            context["datos_liquidos"]["tipos"]["Aceite histórico"],
            40.0,
        )
        self.assertContains(response, 'data-feature-disabled="liquidos"')
        self.assertContains(response, 'id="chartLiquidos"')
        self.assertContains(response, "Contenedor de Líquidos")
        self.assertContains(response, "La lógica de cálculo se conserva")
        self.assertContains(response, 'id="datos_liquidos_json"')
        self.assertNotContains(response, "Vaciar Líquidos")

    def test_dashboard_permite_abrir_un_pendiente_y_explica_que_falta(self):
        pendiente = Residuo.objects.create(
            taller=self.taller,
            seccion=self.taller.seccion,
            profesor=self.taller.profesor,
            asignatura=self.taller.asignatura,
            horario=self.taller.horario,
            numero_clase=1,
            estado=ESTADO_RESIDUO_PENDIENTE,
            created_by=self.user,
        )

        response = self.client.get(reverse("dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Presione para revisar, terminar o cerrar")
        self.assertContains(response, f"Pendiente #{pendiente.id}")
        self.assertContains(response, "Qué hay que hacer")
        self.assertContains(response, "Terminar registro")
        self.assertContains(response, "Cerrar pendiente")

    def test_terminar_pendiente_reanuda_el_mismo_registro(self):
        pendiente = Residuo.objects.create(
            taller=self.taller,
            seccion=self.taller.seccion,
            profesor=self.taller.profesor,
            asignatura=self.taller.asignatura,
            horario=self.taller.horario,
            numero_clase=1,
            estado=ESTADO_RESIDUO_PENDIENTE,
            created_by=self.user,
        )

        response = self.client.get(reverse("residuos"), {"pendiente": pendiente.id})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["pendiente_reanudado"]["residuo_id"], pendiente.id)
        self.assertContains(response, f'"residuo_id": {pendiente.id}')

    def test_cerrar_pendiente_lo_anula_y_lo_retira_del_panel(self):
        pendiente = Residuo.objects.create(
            taller=self.taller,
            seccion=self.taller.seccion,
            profesor=self.taller.profesor,
            asignatura=self.taller.asignatura,
            horario=self.taller.horario,
            numero_clase=1,
            estado=ESTADO_RESIDUO_PENDIENTE,
            created_by=self.user,
        )

        response = self.client.post(reverse("cerrar_residuo_pendiente", args=[pendiente.id]))

        self.assertRedirects(response, reverse("dashboard"))
        pendiente.refresh_from_db()
        self.assertEqual(pendiente.estado, ESTADO_RESIDUO_ANULADO)
        self.assertTrue(pendiente.retirado)
        self.assertEqual(crear_dashboard_context(self.user)["residuos_pendientes"], 0)

    def test_usuario_no_puede_terminar_ni_cerrar_pendiente_ajeno(self):
        otro_usuario = User.objects.create_user(
            username="otro@test.cl",
            password="ClaveSegura123!",
        )
        pendiente = Residuo.objects.create(
            taller=self.taller,
            seccion=self.taller.seccion,
            profesor=self.taller.profesor,
            asignatura=self.taller.asignatura,
            horario=self.taller.horario,
            numero_clase=1,
            estado=ESTADO_RESIDUO_PENDIENTE,
            created_by=otro_usuario,
        )

        dashboard = self.client.get(reverse("dashboard"))
        self.assertContains(dashboard, "Debe terminarlo el usuario que lo inició o un administrador")
        self.assertNotContains(dashboard, "Terminar registro")

        continuar = self.client.get(reverse("residuos"), {"pendiente": pendiente.id})
        self.assertRedirects(continuar, reverse("dashboard"))
        cerrar = self.client.post(reverse("cerrar_residuo_pendiente", args=[pendiente.id]))
        self.assertRedirects(cerrar, reverse("dashboard"))

        pendiente.refresh_from_db()
        self.assertEqual(pendiente.estado, ESTADO_RESIDUO_PENDIENTE)

    def test_guardar_ignora_datos_de_taller_manipulados_por_el_cliente(self):
        response = self.client.post(reverse("guardar"), {
            "taller_id": self.taller.id,
            "asignatura": "Taller falsificado",
            "seccion": "Sección falsificada",
            "profesor": "Usuario falsificado",
            "horario": "01:00 - 02:00",
        })

        self.assertEqual(response.status_code, 200)
        residuo = Residuo.objects.get(pk=response.json()["id"])
        self.assertEqual(residuo.taller, self.taller)
        self.assertEqual(residuo.asignatura, self.taller.asignatura)
        self.assertEqual(residuo.seccion, self.taller.seccion)
        self.assertEqual(residuo.profesor, self.taller.profesor)

    def test_registro_normal_permite_categoria_inorganica_en_unidades(self):
        categoria = CategoriaResiduos.objects.filter(
            tipo_operacional=TIPO_INORGANICO,
        ).order_by("id_categoria").first()
        residuo = Residuo.objects.create(
            seccion="A1",
            profesor="Profesor Demo",
            asignatura="Taller Inorgánico",
            horario="10:00 - 11:00",
            numero_clase=1,
            estado=ESTADO_RESIDUO_PENDIENTE,
            created_by=self.user,
        )

        response = self.client.post(reverse("actualizar_residuo"), {
            "residuo_id": residuo.id,
            "tipo": categoria.id_categoria,
            "unidad": "250",
        })

        self.assertEqual(LIMITE_UNIDADES_GLOBAL, 1000)
        self.assertEqual(response.status_code, 200)
        residuo.refresh_from_db()
        self.assertEqual(residuo.tipo.tipo_operacional, TIPO_INORGANICO)
        self.assertEqual(residuo.unidad, 250)
        self.assertIsNone(residuo.peso)
        self.assertEqual(residuo.contenedor_id, "Unidades")

    def test_unidades_actualizan_total_y_porcentaje_del_grafico(self):
        categoria = CategoriaResiduos.objects.filter(
            tipo_operacional=TIPO_INORGANICO,
        ).order_by("id_categoria").first()
        subtipo = TipoResiduos.objects.get(nombre_residuo="Inorgánico")
        Residuo.objects.create(
            seccion="A1",
            profesor="Profesor Demo",
            asignatura="Taller Inorgánico",
            horario="10:00 - 11:00",
            numero_clase=1,
            tipo=categoria,
            subtipo=subtipo,
            unidad=250,
            contenedor_id="Unidades",
            estado=ESTADO_RESIDUO_CONFIRMADO,
            confirmado_at=timezone.now(),
            created_by=self.user,
        )

        datos = crear_dashboard_context(self.user)["datos_unidades"]

        self.assertEqual(datos["total"], 250)
        self.assertEqual(datos["limite_total"], 1000)
        self.assertEqual(datos["porcentaje_global"], 25.0)
        self.assertEqual(datos["tipos"], {"Inorgánico": 250})
        endpoint = self.client.get(reverse("dashboard_unidades"))
        self.assertEqual(endpoint.status_code, 200)
        self.assertEqual(endpoint.json()["total"], 250)
        dashboard = self.client.get(reverse("dashboard"))
        self.assertContains(dashboard, "250 / 1000")

    def test_limite_de_unidades_considera_el_total_acumulado(self):
        categoria = CategoriaResiduos.objects.filter(
            tipo_operacional=TIPO_INORGANICO,
        ).order_by("id_categoria").first()
        subtipo = TipoResiduos.objects.get(nombre_residuo="Inorgánico")
        Residuo.objects.create(
            seccion="A1",
            profesor="Profesor Demo",
            asignatura="Taller anterior",
            horario="08:00 - 09:00",
            numero_clase=1,
            tipo=categoria,
            subtipo=subtipo,
            unidad=900,
            contenedor_id="Unidades",
            estado=ESTADO_RESIDUO_CONFIRMADO,
            confirmado_at=timezone.now(),
            created_by=self.user,
        )
        pendiente = Residuo.objects.create(
            taller=self.taller,
            seccion=self.taller.seccion,
            profesor=self.taller.profesor,
            asignatura=self.taller.asignatura,
            horario=self.taller.horario,
            numero_clase=1,
            estado=ESTADO_RESIDUO_PENDIENTE,
            created_by=self.user,
        )

        response = self.client.post(reverse("actualizar_residuo"), {
            "residuo_id": pendiente.id,
            "tipo": categoria.id_categoria,
            "unidad": "101",
        })

        self.assertEqual(response.status_code, 400)
        self.assertIn("solo admite 100 adicionales", response.json()["error"])
        self.assertEqual(
            crear_dashboard_context(self.user)["datos_unidades"]["total"],
            900,
        )

    def test_dashboard_no_suma_residuos_pendientes(self):
        Residuo.objects.create(
            seccion="A1",
            profesor="Profesor Demo",
            asignatura="Taller Compost",
            horario="10:00 - 11:00",
            numero_clase=1,
            tipo=self.categoria,
            subtipo=self.subtipo,
            peso=Decimal("2.000"),
            contenedor_id="Compostera 1",
            estado=ESTADO_RESIDUO_CONFIRMADO,
            confirmado_at=timezone.now(),
            created_by=self.user,
        )
        Residuo.objects.create(
            seccion="A1",
            profesor="Profesor Demo",
            asignatura="Taller Compost",
            horario="10:00 - 11:00",
            numero_clase=2,
            tipo=self.categoria,
            subtipo=self.subtipo,
            peso=Decimal("9.000"),
            contenedor_id="Compostera 1",
            estado=ESTADO_RESIDUO_PENDIENTE,
            created_by=self.user,
        )

        context = crear_dashboard_context(self.user)

        self.assertEqual(context["total_peso_organico"], 2.0)
        self.assertEqual(context["residuos_pendientes"], 1)

    def test_asignacion_organica_usa_c1_luego_c2_y_no_c3(self):
        self._crear_residuo_organico("Compostera 1", "75.000", 1)

        self.assertEqual(
            obtener_proximo_contenedor_disponible(Decimal("0.500")),
            "Compostera 2",
        )

        self._crear_residuo_organico("Compostera 2", "75.000", 2)
        self._crear_residuo_organico("Compostera 3", "1.000", 3)

        self.assertIsNone(obtener_proximo_contenedor_disponible(Decimal("0.500")))

    def test_ingreso_de_120_kg_se_reparte_entre_c1_y_c2(self):
        residuo = Residuo.objects.create(
            seccion="A1",
            profesor="Profesor Demo",
            asignatura="Taller Compost",
            horario="10:00 - 11:00",
            numero_clase=1,
            estado=ESTADO_RESIDUO_PENDIENTE,
            created_by=self.user,
        )

        response = self.client.post(reverse("actualizar_residuo"), {
            "residuo_id": residuo.id,
            "tipo": self.categoria.id_categoria,
            "subtipo": self.subtipo.id_tipo,
            "peso": "120.000",
        })

        self.assertEqual(response.status_code, 200)
        asignaciones = list(
            Residuo.objects.filter(estado=ESTADO_RESIDUO_CONFIRMADO)
            .order_by("numero_clase")
            .values_list("contenedor_id", "peso")
        )
        self.assertEqual(asignaciones, [
            ("Compostera 1", Decimal("75.000")),
            ("Compostera 2", Decimal("45.000")),
        ])
        self.assertIn("Compostera 1 (75.000 kg)", response.json()["contenedor_asignado"])
        self.assertIn("Compostera 2 (45.000 kg)", response.json()["contenedor_asignado"])

    def test_rotacion_vuelve_a_c1_cuando_se_vacia(self):
        c1 = self._crear_residuo_organico("Compostera 1", "75.000", 1)
        self._crear_residuo_organico("Compostera 2", "45.000", 2)
        c1.retirado = True
        c1.save(update_fields=["retirado"])

        self.assertEqual(
            obtener_proximo_contenedor_disponible(Decimal("10.000")),
            "Compostera 1",
        )

    def test_cambio_de_c1_a_c2_no_duplica_registro(self):
        self._crear_residuo_organico("Compostera 1", "75.000", 1)
        residuo = Residuo.objects.create(
            seccion="A1",
            profesor="Profesor Demo",
            asignatura="Taller Compost",
            horario="10:00 - 11:00",
            numero_clase=2,
            estado=ESTADO_RESIDUO_PENDIENTE,
            created_by=self.user,
        )
        total_antes = Residuo.objects.count()

        response = self.client.post(reverse("actualizar_residuo"), {
            "residuo_id": residuo.id,
            "tipo": self.categoria.id_categoria,
            "subtipo": self.subtipo.id_tipo,
            "peso": "0.500",
        })

        self.assertEqual(response.status_code, 200)
        residuo.refresh_from_db()
        self.assertEqual(residuo.contenedor_id, "Compostera 2")
        self.assertEqual(Residuo.objects.count(), total_antes)

    def test_no_asigna_c3_si_c1_y_c2_estan_llenas(self):
        self._crear_residuo_organico("Compostera 1", "75.000", 1)
        self._crear_residuo_organico("Compostera 2", "75.000", 2)
        residuo = Residuo.objects.create(
            seccion="A1",
            profesor="Profesor Demo",
            asignatura="Taller Compost",
            horario="10:00 - 11:00",
            numero_clase=3,
            estado=ESTADO_RESIDUO_PENDIENTE,
            created_by=self.user,
        )

        response = self.client.post(reverse("actualizar_residuo"), {
            "residuo_id": residuo.id,
            "tipo": self.categoria.id_categoria,
            "subtipo": self.subtipo.id_tipo,
            "peso": "0.500",
        })

        self.assertEqual(response.status_code, 400)
        self.assertIn("Compostera 3 está deshabilitada", response.json()["error"])
        residuo.refresh_from_db()
        self.assertNotEqual(residuo.contenedor_id, "Compostera 3")


class CapacidadUnidadesConcurrenteTests(TransactionTestCase):
    reset_sequences = True

    def setUp(self):
        OperationalLock.objects.get_or_create(name="unidades")
        grupo, _ = Group.objects.get_or_create(name="Estudiante")
        self.users = []
        self.clients = []
        for indice in (1, 2):
            user = User.objects.create_user(
                username=f"concurrente{indice}@test.cl",
                password="ClaveSegura123!",
            )
            user.groups.add(grupo)
            client = Client()
            client.force_login(user)
            self.users.append(user)
            self.clients.append(client)

        self.categoria = CategoriaResiduos.objects.create(
            nombre="Inorgánico concurrente",
            descripcion="Prueba de concurrencia",
            tipo_operacional=TIPO_INORGANICO,
        )
        self.subtipo = TipoResiduos.objects.create(
            categoria=self.categoria,
            nombre_residuo="Inorgánico concurrente",
            descripcion="Prueba",
        )
        Residuo.objects.create(
            seccion="A1",
            profesor="Profesor",
            asignatura="Acumulado previo",
            horario="08:00 - 09:00",
            numero_clase=1,
            tipo=self.categoria,
            subtipo=self.subtipo,
            unidad=400,
            contenedor_id="Unidades",
            estado=ESTADO_RESIDUO_CONFIRMADO,
            confirmado_at=timezone.now(),
            created_by=self.users[0],
        )
        self.pendientes = [
            Residuo.objects.create(
                seccion="A1",
                profesor="Profesor",
                asignatura=f"Registro concurrente {indice}",
                horario="09:00 - 10:00",
                numero_clase=1,
                estado=ESTADO_RESIDUO_PENDIENTE,
                created_by=user,
            )
            for indice, user in enumerate(self.users, start=1)
        ]

    def test_dos_usuarios_no_pueden_superar_el_limite_acumulado(self):
        barrera = threading.Barrier(2)
        resultados = [None, None]

        def registrar(indice):
            close_old_connections()
            barrera.wait(timeout=5)
            try:
                with transaction.atomic():
                    _bloquear_capacidad("unidades")
                    _validar_capacidad_unidades(400)
                    Residuo.objects.create(
                        seccion="A1",
                        profesor="Profesor",
                        asignatura=f"Confirmado concurrente {indice}",
                        horario="09:00 - 10:00",
                        numero_clase=1,
                        tipo=self.categoria,
                        subtipo=self.subtipo,
                        unidad=400,
                        contenedor_id="Unidades",
                        estado=ESTADO_RESIDUO_CONFIRMADO,
                        confirmado_at=timezone.now(),
                        created_by=self.users[indice],
                    )
                resultados[indice] = 200
            except ValueError:
                resultados[indice] = 400
            except OperationalError:
                resultados[indice] = 409
            close_old_connections()

        hilos = [threading.Thread(target=registrar, args=(indice,)) for indice in (0, 1)]
        for hilo in hilos:
            hilo.start()
        for hilo in hilos:
            hilo.join(timeout=10)

        total = (
            Residuo.objects.filter(
                estado=ESTADO_RESIDUO_CONFIRMADO,
                retirado=False,
                tipo__tipo_operacional=TIPO_INORGANICO,
            ).aggregate(total=Sum("unidad"))["total"] or 0
        )
        self.assertLessEqual(total, 1000)
        self.assertTrue(all(status in {200, 400, 409} for status in resultados))
        self.assertTrue(any(status in {400, 409} for status in resultados))


class WeightApiTests(TestCase):
    def setUp(self):
        grupo, _ = Group.objects.get_or_create(name="Estudiante")
        self.user = User.objects.create_user(
            username="operador@test.cl",
            email="operador@test.cl",
            password="ClaveSegura123!",
        )
        self.user.groups.add(grupo)
        self.client.force_login(self.user)

    @override_settings(WEIGHT_API_TOKEN="token-demo")
    def test_api_balanza_requiere_token_y_registra_peso(self):
        body = {
            "weight_kg": "3.456",
            "device_name": "balanza-test",
            "raw_data": "3.456 kg",
            "is_stable": True,
        }

        response = self.client.post(
            reverse("api_update_weight"),
            data=json.dumps(body),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 403)

        response = self.client.post(
            reverse("api_update_weight"),
            data=json.dumps(body),
            content_type="application/json",
            HTTP_X_WEIGHT_TOKEN="token-demo",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(WeightReading.objects.count(), 1)
        lectura = WeightReading.objects.get()
        self.assertEqual(lectura.weight_kg, Decimal("3.456"))
        self.assertTrue(lectura.is_stable)

    def test_last_weight_entrega_ultima_lectura_estable(self):
        WeightReading.objects.create(weight_kg=Decimal("1.000"), device_name="estable", is_stable=True)
        WeightReading.objects.create(weight_kg=Decimal("9.000"), device_name="inestable", is_stable=False)

        response = self.client.get(reverse("last_weight"))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["weight_kg"], 1.0)
        self.assertEqual(payload["device_name"], "estable")

    @override_settings(WEIGHT_READING_MAX_AGE_SECONDS=10)
    def test_last_weight_no_reutiliza_una_lectura_antigua(self):
        reading = WeightReading.objects.create(
            weight_kg=Decimal("1.000"),
            device_name="antigua",
            is_stable=True,
        )
        WeightReading.objects.filter(pk=reading.pk).update(
            created_at=timezone.now() - timedelta(seconds=11),
        )

        response = self.client.get(reverse("last_weight"))

        self.assertEqual(response.status_code, 200)
        self.assertIsNone(response.json()["weight_kg"])

    def test_formulario_no_incluye_simulador_de_peso(self):
        response = self.client.get(reverse("residuos"))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Simular lectura")
        self.assertContains(response, "Buscar balanza automáticamente")

    def test_normalizar_peso_convierte_gramos_a_kg(self):
        self.assertEqual(
            balanza_service.normalizar_peso("ST,GS,+700 g"),
            Decimal("0.700"),
        )

    def test_normalizar_peso_ignora_trama_serial_corrupta(self):
        self.assertIsNone(balanza_service.normalizar_peso("\x08HRo7)\x00"))

    def test_normalizar_peso_lee_trama_real_de_la_balanza(self):
        self.assertEqual(
            balanza_service.normalizar_peso("\x02S  0.245kgr\x03"),
            Decimal("0.245"),
        )

    @override_settings(
        WEIGHT_BRIDGE_FIRST=False,
        WEIGHT_DIRECT_READ_ENABLED=True,
    )
    @patch("app.views.balanza_service.obtener_peso_estable")
    def test_balanza_leer_registra_lectura_automatica(self, mock_peso):
        mock_peso.return_value = {
            "ok": True,
            "peso": Decimal("2.350"),
            "unidad": "kg",
            "puerto": "COM5",
            "baudrate": 9600,
            "serial_mode": "8N1",
            "line_control": "default",
            "mensaje": "Lectura correcta",
            "raw": "2.350 kg",
            "dispositivo": "USB-SERIAL (COM5)",
            "fuente": "serial",
            "puertos_probados": ["COM5"],
        }

        response = self.client.get(reverse("balanza_leer"))

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["ok"])
        self.assertEqual(response.json()["peso"], 2.35)
        self.assertEqual(WeightReading.objects.get().weight_kg, Decimal("2.350"))

    @override_settings(
        BALANZA_SERIAL_BAUDRATES=["9600"],
        BALANZA_SERIAL_MODES=["8N1"],
        BALANZA_LINE_CONTROLS=["default"],
        BALANZA_SCAN_TIMEOUT_SECONDS=20,
    )
    @patch("app.services.balanza_service._leer_usb_estable")
    @patch("app.services.balanza_service._leer_hid_estable")
    @patch("app.services.balanza_service._leer_serial_estable")
    @patch("app.services.balanza_service.listar_dispositivos_balanza")
    def test_prueba_todos_los_com_y_luego_hid_usb(
        self,
        mock_dispositivos,
        mock_serial,
        mock_hid,
        mock_usb,
    ):
        mock_dispositivos.return_value = {
            "serial": [
                {"puerto": "COM8", "configurado": False, "probable_balanza": False},
                {"puerto": "COM7", "configurado": False, "probable_balanza": False},
            ],
            "hid": [{"path": b"hid-test", "probable_balanza": True}],
            "usb": [{"vendor_id": 1, "product_id": 2, "probable_balanza": True}],
            "errores": [],
        }
        mock_serial.return_value = {
            "ok": False,
            "mensaje": "sin lectura",
            "errores": [],
        }
        mock_hid.return_value = {
            "ok": False,
            "mensaje": "sin lectura HID",
            "errores": [],
        }
        mock_usb.return_value = {
            "ok": False,
            "mensaje": "sin lectura USB",
            "errores": [],
        }

        resultado = balanza_service._obtener_peso_estable_sin_bloqueo()

        self.assertFalse(resultado["ok"])
        self.assertEqual(
            [call.args[0] for call in mock_serial.call_args_list],
            ["COM7", "COM8"],
        )
        mock_hid.assert_called_once()
        mock_usb.assert_called_once()
        self.assertEqual(resultado["puertos_probados"], ["COM7", "COM8"])


class HorariosProfesorTests(TestCase):
    def setUp(self):
        grupo, _ = Group.objects.get_or_create(name=ROL_PROFESOR)
        self.user = User.objects.create_user(
            username="profesor@test.cl",
            email="profesor@test.cl",
            password="ClaveSegura123!",
        )
        self.user.groups.add(grupo)
        self.client.force_login(self.user)

    def test_profesor_muestra_fecha_del_horario(self):
        fecha = timezone.localdate()
        ClaseHorario.objects.create(
            profesor=self.user.username,
            asignatura="Aceite",
            seccion="Medicion aceite",
            horario="12:46 - 12:48",
            fecha=fecha,
            dia_semana=fecha.weekday(),
            hora_inicio=time(12, 46),
            hora_fin=time(12, 48),
        )

        response = self.client.get(reverse("profesor"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, fecha.strftime("%d-%m-%Y"))

    def test_archivar_pasadas_incluye_actividades_de_hoy_terminadas(self):
        ahora = timezone.make_aware(datetime(2026, 6, 26, 13, 0))
        fecha = ahora.date()
        pasada = ClaseHorario.objects.create(
            profesor=self.user.username,
            asignatura="Aceite",
            seccion="Medicion aceite",
            horario="12:00 - 12:30",
            fecha=fecha,
            dia_semana=fecha.weekday(),
            hora_inicio=time(12, 0),
            hora_fin=time(12, 30),
        )
        vigente = ClaseHorario.objects.create(
            profesor=self.user.username,
            asignatura="Taller",
            seccion="Vigente",
            horario="13:00 - 14:00",
            fecha=fecha,
            dia_semana=fecha.weekday(),
            hora_inicio=time(13, 0),
            hora_fin=time(14, 0),
        )

        with patch("app.views.timezone.localtime", return_value=ahora):
            response = self.client.post(reverse("archivar_actividades_pasadas"))
            self.assertRedirects(response, reverse("profesor"))

        pasada.refresh_from_db()
        vigente.refresh_from_db()
        self.assertTrue(pasada.archivado)
        self.assertFalse(vigente.archivado)

    def test_profesor_puede_eliminar_su_actividad(self):
        fecha = timezone.localdate()
        clase = ClaseHorario.objects.create(
            profesor=self.user.username,
            asignatura="Aceite",
            seccion="Medicion aceite",
            horario="12:46 - 12:48",
            fecha=fecha,
            dia_semana=fecha.weekday(),
            hora_inicio=time(12, 46),
            hora_fin=time(12, 48),
        )

        response = self.client.post(reverse("eliminar_actividad", args=[clase.id]))

        self.assertRedirects(response, reverse("profesor"))
        self.assertFalse(ClaseHorario.objects.filter(pk=clase.pk).exists())

    def test_profesor_no_puede_eliminar_actividad_de_otro_profesor(self):
        fecha = timezone.localdate()
        clase = ClaseHorario.objects.create(
            profesor="otro@test.cl",
            asignatura="Aceite",
            seccion="Medicion aceite",
            horario="12:46 - 12:48",
            fecha=fecha,
            dia_semana=fecha.weekday(),
            hora_inicio=time(12, 46),
            hora_fin=time(12, 48),
        )

        response = self.client.post(reverse("eliminar_actividad", args=[clase.id]))

        self.assertEqual(response.status_code, 404)
        self.assertTrue(ClaseHorario.objects.filter(pk=clase.pk).exists())

    def test_boton_archivar_no_queda_deshabilitado_si_no_hay_pasadas(self):
        fecha = timezone.localdate()
        ClaseHorario.objects.create(
            profesor=self.user.username,
            asignatura="Taller futuro",
            seccion="Medicion",
            horario="23:00 - 23:30",
            fecha=fecha,
            dia_semana=fecha.weekday(),
            hora_inicio=time(23, 0),
            hora_fin=time(23, 30),
        )

        response = self.client.get(reverse("profesor"))

        self.assertContains(response, "Archivar talleres finalizados")
        contenido = response.content.decode()
        texto_boton = contenido.index("Archivar talleres finalizados")
        etiqueta_boton = contenido[
            contenido.rfind("<button", 0, texto_boton):contenido.find(">", texto_boton)
        ]
        self.assertNotIn("disabled", etiqueta_boton)

    def test_creacion_de_taller_muestra_cuatro_opciones_antes_del_formulario(self):
        response = self.client.get(reverse("profesor"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Crear taller", count=1)
        self.assertEqual(
            list(response.context["tipo_taller_opciones"]),
            [
                (ClaseHorario.TIPO_RECUPERACION_CLASES, "Recuperación de clases"),
                (ClaseHorario.TIPO_SESION_EXTRA, "Sesión extra"),
                (
                    ClaseHorario.TIPO_ACTIVIDAD_EXTRA_PROGRAMATICA,
                    "Actividad extra programática",
                ),
                (ClaseHorario.TIPO_NUEVA_ACTIVIDAD, "Nueva actividad"),
            ],
        )
        contenido = response.content.decode()
        self.assertContains(response, "Selecciona una opción")
        self.assertContains(response, 'class="tipo-taller-option"', count=4, html=False)
        self.assertContains(response, "Recuperación de clases")
        self.assertContains(response, "Sesión extra")
        self.assertContains(response, "Actividad extra programática")
        self.assertContains(response, "Nueva actividad")
        self.assertRegex(contenido, r'id="taller-creation-options"[^>]*hidden')
        self.assertRegex(contenido, r'id="taller-form-panel"[^>]*hidden')
        self.assertIn('id="id_tipo_taller"', contenido)
        self.assertIn('data-tipo="recuperacion_clases"', contenido)
        self.assertIn('data-tipo="sesion_extra"', contenido)
        self.assertIn('data-tipo="actividad_extra_programatica"', contenido)
        self.assertIn('data-tipo="nueva_actividad"', contenido)
        self.assertIn("taller_precargado", contenido)

    def test_footer_muestra_los_ocho_enlaces_sociales_con_imagenes(self):
        response = self.client.get(reverse("profesor"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'class="social-link social-link--', count=8, html=False)
        self.assertContains(response, 'class="social-icon"', count=6, html=False)
        self.assertContains(response, 'class="social-icon social-icon--stroke"', count=2, html=False)
        self.assertContains(response, 'class="social-name"', count=8, html=False)
        self.assertContains(
            response,
            'href="https://www.linkedin.com/school/inacap/"',
            html=False,
        )
        self.assertContains(
            response,
            'href="https://www.instagram.com/inacap_oficial/"',
            html=False,
        )
        self.assertContains(
            response,
            'href="https://www.tiktok.com/@inacap_oficial?lang=es"',
            html=False,
        )
        self.assertContains(response, 'href="https://www.facebook.com/INACAP"', html=False)
        self.assertContains(
            response,
            'href="https://www.youtube.com/user/CANALINACAP"',
            html=False,
        )
        self.assertContains(response, 'href="https://x.com/INACAPINOS"', html=False)
        self.assertContains(
            response,
            'href="https://www.youtube.com/playlist?list=PL3EFFi1Jpv7sLv86EkdKiOix6JDrE34sJ"',
            html=False,
        )
        self.assertContains(
            response,
            'href="https://open.spotify.com/show/3jgaLBiDwakcWY44WKd7DD?si=e9703b079c614bdc"',
            html=False,
        )

    def test_codigo_qr_muestra_titulo_institucional_superior(self):
        response = self.client.get(reverse("codigo_qr"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            '<h1 class="page-title-banner qr-page-title">QR para registrar residuos</h1>',
            html=True,
        )

    def test_inicio_de_sesion_no_muestra_acceso_qr_de_profesor(self):
        response = Client().get(reverse("login"))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Acceso profesor")
        self.assertNotContains(response, reverse("generar_qr_profesor"))
        self.assertContains(response, "login.css", count=1)

    def test_muestra_los_tres_talleres_precargados(self):
        ahora = timezone.make_aware(datetime(2026, 7, 8, 14, 0))

        with patch("app.views.timezone.localtime", return_value=ahora):
            response = self.client.get(reverse("profesor"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            [(taller["nombre"], taller["activo"]) for taller in response.context["talleres_opciones"]],
            [
                ("Taller 1", False),
                ("Taller 2", True),
                ("Taller 3", False),
            ],
        )
        self.assertContains(response, ">Taller 1</td>", html=False)
        self.assertContains(response, ">Taller 2</td>", html=False)
        self.assertContains(response, ">Taller 3</td>", html=False)

    def test_muestra_talleres_precargados_en_fin_de_semana(self):
        ahora = timezone.make_aware(datetime(2026, 7, 11, 14, 0))

        with patch("app.views.timezone.localtime", return_value=ahora):
            response = self.client.get(reverse("profesor"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            [taller["nombre"] for taller in response.context["talleres_opciones"]],
            ["Taller 1", "Taller 2", "Taller 3"],
        )
        self.assertContains(response, "Talleres precargados")
        self.assertContains(response, 'class="cargar-taller-form"', count=3, html=False)

    def test_cargar_taller_precargado_redirige_a_residuos_y_no_lo_duplica(self):
        ahora = timezone.make_aware(datetime(2026, 7, 8, 14, 0))

        with patch("app.views.timezone.localtime", return_value=ahora):
            response = self.client.post(
                reverse("profesor"),
                {"taller_precargado": "Taller 2"},
            )
            segunda_respuesta = self.client.post(
                reverse("profesor"),
                {"taller_precargado": "Taller 2"},
            )
            self.assertRedirects(response, reverse("registrar_evento"))
            self.assertRedirects(segunda_respuesta, reverse("registrar_evento"))

        talleres = ClaseHorario.objects.filter(
            profesor=self.user.username,
            fecha=ahora.date(),
            asignatura="Taller 2",
            archivado=False,
        )
        self.assertEqual(talleres.count(), 1)
        self.assertEqual(talleres.get().hora_inicio, time(13, 40))

    def test_crear_taller_registra_el_tipo_seleccionado(self):
        fecha = timezone.localdate() + timedelta(days=1)
        response = self.client.post(
            reverse("profesor"),
            {
                "tipo_taller": ClaseHorario.TIPO_SESION_EXTRA,
                "fecha": fecha.isoformat(),
                "asignatura": "Laboratorio extraordinario",
                "hora_inicio": "10:00",
                "hora_fin": "11:30",
                "seccion": "Sección A",
            },
        )

        self.assertRedirects(response, reverse("profesor"))
        taller = ClaseHorario.objects.get(
            profesor=self.user.username,
            asignatura="Laboratorio extraordinario",
        )
        self.assertEqual(taller.tipo_taller, ClaseHorario.TIPO_SESION_EXTRA)
        self.assertEqual(taller.get_tipo_taller_display(), "Sesión extra")

    def test_no_crea_taller_sin_seleccionar_tipo(self):
        fecha = timezone.localdate() + timedelta(days=1)
        response = self.client.post(
            reverse("profesor"),
            {
                "fecha": fecha.isoformat(),
                "asignatura": "Taller sin tipo",
                "hora_inicio": "10:00",
                "hora_fin": "11:30",
                "seccion": "Sección A",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn("tipo_taller", response.context["form"].errors)
        self.assertFalse(
            ClaseHorario.objects.filter(
                profesor=self.user.username,
                asignatura="Taller sin tipo",
            ).exists()
        )

    def test_administrador_puede_crear_taller(self):
        grupo_admin, _ = Group.objects.get_or_create(name=ROL_ADMINISTRADOR)
        admin = User.objects.create_user(
            username="admin-taller@test.cl",
            email="admin-taller@test.cl",
            password="ClaveSegura123!",
        )
        admin.groups.add(grupo_admin)
        self.client.force_login(admin)
        fecha = timezone.localdate() + timedelta(days=1)

        response = self.client.post(
            reverse("profesor"),
            {
                "tipo_taller": ClaseHorario.TIPO_NUEVA_ACTIVIDAD,
                "fecha": fecha.isoformat(),
                "asignatura": "Taller administrativo",
                "hora_inicio": "10:00",
                "hora_fin": "11:30",
                "seccion": "Sección A",
            },
        )

        self.assertRedirects(response, reverse("profesor"))
        self.assertTrue(
            ClaseHorario.objects.filter(
                asignatura="Taller administrativo",
                profesor=admin.username,
                tipo_taller=ClaseHorario.TIPO_NUEVA_ACTIVIDAD,
            ).exists()
        )

    def test_navegacion_usa_los_nombres_actualizados(self):
        response = self.client.get(reverse("dashboard"))
        contenido = response.content.decode()

        self.assertContains(response, "Panel de Sostenibilidad")
        self.assertIn(">Taller</a>", contenido)
        self.assertNotIn(">Dashboard</a>", contenido)
        self.assertNotIn(">Actividad</a>", contenido)

    def test_formulario_rechaza_hora_fin_anterior_a_inicio(self):
        form = ClaseHorarioForm(data={
            "tipo_taller": ClaseHorario.TIPO_NUEVA_ACTIVIDAD,
            "fecha": timezone.localdate().isoformat(),
            "asignatura": "Aceite",
            "seccion": "Medicion aceite",
            "hora_inicio": "23:27",
            "hora_fin": "02:31",
        })

        self.assertFalse(form.is_valid())
        self.assertIn("La hora de fin debe ser posterior", str(form.errors))


class ExportacionAutomaticaTests(TestCase):
    def setUp(self):
        self.export_dir = TemporaryDirectory()
        self.addCleanup(self.export_dir.cleanup)
        self.export_settings = override_settings(
            EXPORT_DIR=Path(self.export_dir.name)
        )
        self.export_settings.enable()
        self.addCleanup(self.export_settings.disable)

        grupo, _ = Group.objects.get_or_create(name=ROL_PROFESOR)
        self.user = User.objects.create_user(
            username="profesor-export@test.cl",
            email="profesor-export@test.cl",
            password="ClaveSegura123!",
        )
        self.user.groups.add(grupo)
        self.client.force_login(self.user)

        self.categoria = CategoriaResiduos.objects.create(
            nombre="Orgánico",
            descripcion="Residuos compostables",
            tipo_operacional=TIPO_ORGANICO,
        )
        self.subtipo = TipoResiduos.objects.create(
            categoria=self.categoria,
            nombre_residuo="Restos vegetales",
            descripcion="Restos de cocina",
        )

    def _crear_taller(self, ahora, nombre="Taller Export"):
        fecha = ahora.date()
        return ClaseHorario.objects.create(
            profesor=self.user.username,
            asignatura=nombre,
            seccion="Seccion Export",
            horario="10:00 - 11:00",
            fecha=fecha,
            dia_semana=fecha.weekday(),
            hora_inicio=time(10, 0),
            hora_fin=time(11, 0),
        )

    def _crear_residuo_taller(self, clase, nombre=None, numero=1, motivo=None):
        return Residuo.objects.create(
            seccion=clase.seccion,
            profesor=clase.profesor,
            asignatura=nombre or clase.asignatura,
            horario=clase.horario,
            numero_clase=numero,
            hora_escaneo=timezone.make_aware(datetime.combine(clase.fecha, time(10, 30))),
            tipo=self.categoria,
            subtipo=self.subtipo,
            peso=Decimal("1.000"),
            contenedor_id="Compostera 1",
            estado=ESTADO_RESIDUO_CONFIRMADO,
            confirmado_at=timezone.now(),
            created_by=self.user,
            manual_entry=bool(motivo),
            manual_reason=motivo,
        )

    def _worksheet_text(self, path):
        with ZipFile(path) as archive:
            return archive.read("xl/worksheets/sheet1.xml").decode("utf-8")

    def test_cierre_automatico_genera_archivos_del_taller_sin_duplicar(self):
        ahora = timezone.make_aware(datetime(2026, 7, 10, 12, 0))
        clase = self._crear_taller(ahora)
        self._crear_residuo_taller(clase)
        self._crear_residuo_taller(clase, nombre="Otro Taller", numero=2)

        with TemporaryDirectory() as tmp_dir, override_settings(EXPORT_DIR=Path(tmp_dir)):
            with patch("app.views.timezone.localtime", return_value=ahora):
                response = self.client.get(reverse("dashboard"))

            self.assertEqual(response.status_code, 200)
            clase.refresh_from_db()
            self.assertTrue(clase.archivado)
            self.assertTrue(clase.exportado)
            self.assertTrue(clase.export_residuos_path.endswith("exportar residuos.xlsx"))
            self.assertTrue(clase.export_retiros_path.endswith("exportar retiros.xlsx"))

            archivos = list(Path(tmp_dir).rglob("*.xlsx"))
            self.assertEqual(len(archivos), 2)
            residuos_xml = self._worksheet_text(Path(tmp_dir) / clase.export_residuos_path)
            self.assertIn("Taller Export", residuos_xml)
            self.assertNotIn("Otro Taller", residuos_xml)
            with (Path(tmp_dir) / clase.export_residuos_path).open("rb") as archivo:
                worksheet = load_workbook(BytesIO(archivo.read()), read_only=True, data_only=True).active
            self.assertEqual(worksheet["B3"].value, 1)
            self.assertEqual(worksheet.max_column, 21)
            self.assertEqual(worksheet["A4"].value, "ID")
            self.assertEqual(worksheet["A5"].value, 1)
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
                self.assertTrue(
                    any(cell.value not in (None, "") for cell in row),
                    "La exportación no debe contener filas completamente vacías.",
                )
            for row in worksheet.iter_rows(
                min_row=5,
                max_row=worksheet.max_row,
                min_col=1,
                max_col=worksheet.max_column,
            ):
                self.assertTrue(
                    all(cell.value not in (None, "") for cell in row),
                    "Las filas de datos no deben contener celdas vacías.",
                )

            exportado_at = clase.exportado_at
            with patch("app.views.timezone.localtime", return_value=ahora + timedelta(minutes=5)):
                self.client.get(reverse("dashboard"))
            clase.refresh_from_db()
            self.assertEqual(clase.exportado_at, exportado_at)
            self.assertEqual(len(list(Path(tmp_dir).rglob("*.xlsx"))), 2)

    def test_worker_guarda_excel_sin_descarga_del_navegador(self):
        ahora = timezone.make_aware(datetime(2026, 7, 10, 12, 0))
        clase = self._crear_taller(ahora)
        self._crear_residuo_taller(clase)

        with TemporaryDirectory() as tmp_dir, override_settings(EXPORT_DIR=Path(tmp_dir)):
            with patch("app.views.timezone.localtime", return_value=ahora):
                call_command("workshop_worker", "--once")

            clase.refresh_from_db()
            self.assertTrue(clase.archivado)
            self.assertTrue(clase.exportado)
            self.assertEqual(len(list(Path(tmp_dir).rglob("*.xlsx"))), 2)
            self.assertIn("taller-export-seccion-export", clase.export_residuos_path)
            self.assertIsNone(clase.export_residuos_descargado_at)
            self.assertIsNone(clase.export_retiros_descargado_at)
            base = self.client.get(reverse("dashboard")).content.decode("utf-8")
            self.assertNotIn("exportaciones/pendientes", base)
            self.assertNotIn('link.download', base)

    def test_exportaciones_no_muestran_botones(self):
        ahora = timezone.make_aware(datetime(2026, 7, 10, 12, 0))
        self._crear_taller(ahora)

        with TemporaryDirectory() as tmp_dir, override_settings(EXPORT_DIR=Path(tmp_dir)):
            with patch("app.views.timezone.localtime", return_value=ahora):
                dashboard_response = self.client.get(reverse("dashboard"))
                profesor_response = self.client.get(reverse("profesor"))

        for response in (dashboard_response, profesor_response):
            self.assertNotContains(response, "Descargar residuos")
            self.assertNotContains(response, "Descargar retiros")
            self.assertNotContains(response, "Exportaciones automáticas")

    def test_logout_cierra_taller_y_genera_archivos(self):
        ahora = timezone.make_aware(datetime(2026, 7, 10, 12, 0))
        clase = self._crear_taller(ahora)

        with TemporaryDirectory() as tmp_dir, override_settings(EXPORT_DIR=Path(tmp_dir)):
            with patch("app.views.timezone.localtime", return_value=ahora):
                response = self.client.post(reverse("logout"))

            self.assertRedirects(response, reverse("login"))
            clase.refresh_from_db()
            self.assertTrue(clase.archivado)
            self.assertTrue(clase.exportado)
            self.assertEqual(len(list(Path(tmp_dir).rglob("*.xlsx"))), 2)

    def test_error_de_exportacion_queda_registrado_y_se_puede_reintentar(self):
        ahora = timezone.make_aware(datetime(2026, 7, 10, 12, 0))
        clase = self._crear_taller(ahora)

        with TemporaryDirectory() as tmp_dir, override_settings(EXPORT_DIR=Path(tmp_dir)):
            with patch("app.views.timezone.localtime", return_value=ahora), patch(
                "app.views._guardar_reporte_taller",
                side_effect=RuntimeError("fallo controlado"),
            ):
                response = self.client.get(reverse("dashboard"))

            self.assertEqual(response.status_code, 200)
            clase.refresh_from_db()
            self.assertTrue(clase.archivado)
            self.assertFalse(clase.exportado)
            self.assertIn("fallo controlado", clase.export_error)

            response = self.client.post(reverse("reintentar_exportacion_taller", args=[clase.id]))
            self.assertRedirects(response, reverse("profesor"))
            clase.refresh_from_db()
            self.assertTrue(clase.exportado)
            self.assertEqual(clase.export_error, "")

    @patch("app.views.create_rolling_backup")
    def test_ingreso_manual_registra_trazabilidad_y_actualiza_exportacion(
        self,
        respaldo_mock,
    ):
        respaldo_mock.return_value = Path("respaldo_base_datos.sqlite3")
        ahora = timezone.make_aware(datetime(2026, 7, 10, 12, 0))
        clase = self._crear_taller(ahora)
        clase.archivado = True
        clase.save(update_fields=["archivado"])

        with patch(
            "app.views.is_sqlite_memory_database",
            return_value=False,
        ), TemporaryDirectory() as tmp_dir, override_settings(EXPORT_DIR=Path(tmp_dir)):
            formulario = self.client.get(reverse("ingreso_manual_olvido"))
            self.assertNotContains(formulario, "Subtipo")
            self.assertNotContains(formulario, 'id="id_subtipo"')

            response = self.client.post(reverse("ingreso_manual_olvido"), {
                "fecha": clase.fecha.isoformat(),
                "taller": clase.id,
                "categoria": self.categoria.id_categoria,
                "subtipo": self.subtipo.id_tipo,
                "tipo_medicion": "peso_kg",
                "cantidad": "1.250",
                "motivo": "Registro olvidado durante el taller",
            })

            self.assertRedirects(response, reverse("dashboard"))
            residuo = Residuo.objects.get(manual_entry=True)
            self.assertEqual(residuo.manual_reason, "Registro olvidado durante el taller")
            self.assertEqual(residuo.created_by, self.user)
            self.assertIsNotNone(residuo.created_at)
            self.assertEqual(residuo.peso, Decimal("1.250"))
            self.assertIsNone(residuo.unidad)

            clase.refresh_from_db()
            self.assertTrue(clase.exportado)
            residuos_xml = self._worksheet_text(Path(tmp_dir) / clase.export_residuos_path)
            self.assertIn("Manual", residuos_xml)
            self.assertIn("Registro olvidado durante el taller", residuos_xml)
            respaldo_mock.assert_called_once_with()

    @patch("app.views.create_rolling_backup")
    def test_respaldo_se_actualiza_desde_30_segundos_antes_del_cierre(
        self,
        respaldo_mock,
    ):
        respaldo_mock.return_value = Path("respaldo_base_datos.sqlite3")
        ahora = timezone.make_aware(datetime(2026, 7, 10, 10, 59, 29))
        clase = self._crear_taller(ahora)
        request = RequestFactory().get("/")
        request.user = self.user

        with patch("app.views.is_sqlite_memory_database", return_value=False):
            with patch("app.views.timezone.localtime", return_value=ahora):
                self.assertIsNone(
                    _respaldar_clases_por_finalizar(
                        request,
                        ClaseHorario.objects.filter(pk=clase.pk),
                        ahora=ahora,
                    )
                )
            respaldo_mock.assert_not_called()

            momento_respaldo = ahora + timedelta(seconds=1)
            with patch("app.views.timezone.localtime", return_value=momento_respaldo):
                _respaldar_clases_por_finalizar(
                    request,
                    ClaseHorario.objects.filter(pk=clase.pk),
                    ahora=momento_respaldo,
                )

            clase.refresh_from_db()
            self.assertEqual(clase.respaldo_pre_cierre_at, momento_respaldo)
            respaldo_mock.assert_called_once_with()

    def test_ingreso_manual_grande_se_distribuye_sin_informar_lleno(self):
        ahora = timezone.make_aware(datetime(2026, 7, 10, 12, 0))
        clase = self._crear_taller(ahora, nombre="Taller Distribución")

        response = self.client.post(reverse("ingreso_manual_olvido"), {
            "fecha": clase.fecha.isoformat(),
            "taller": clase.id,
            "categoria": self.categoria.id_categoria,
            "subtipo": self.subtipo.id_tipo,
            "tipo_medicion": "peso_kg",
            "cantidad": "120.000",
            "motivo": "Ingreso grande de prueba",
        })

        self.assertRedirects(response, reverse("dashboard"))
        asignaciones = list(
            Residuo.objects.filter(manual_reason="Ingreso grande de prueba")
            .order_by("numero_clase")
            .values_list("contenedor_id", "peso")
        )
        self.assertEqual(asignaciones, [
            ("Compostera 1", Decimal("75.000")),
            ("Compostera 2", Decimal("45.000")),
        ])

    def test_ingreso_manual_permite_medicion_en_unidades(self):
        ahora = timezone.make_aware(datetime(2026, 7, 10, 12, 0))
        clase = self._crear_taller(ahora, nombre="Taller Unidades")
        categoria = CategoriaResiduos.objects.create(
            nombre="Inorgánico prueba manual",
            descripcion="Materiales por unidad",
            tipo_operacional=TIPO_INORGANICO,
        )
        subtipo = TipoResiduos.objects.create(
            categoria=categoria,
            nombre_residuo="Envases prueba manual",
            descripcion="Envases descartados",
        )

        response = self.client.post(reverse("ingreso_manual_olvido"), {
            "fecha": clase.fecha.isoformat(),
            "taller": clase.id,
            "categoria": categoria.id_categoria,
            "subtipo": subtipo.id_tipo,
            "tipo_medicion": "unidades",
            "cantidad": "7",
            "motivo": "Conteo manual olvidado",
        })

        self.assertRedirects(response, reverse("dashboard"))
        residuo = Residuo.objects.get(manual_reason="Conteo manual olvidado")
        self.assertEqual(residuo.unidad, 7)
        self.assertIsNone(residuo.peso)

    def test_ingreso_manual_respeta_capacidad_acumulada_de_unidades(self):
        ahora = timezone.make_aware(datetime(2026, 7, 10, 12, 0))
        clase = self._crear_taller(ahora, nombre="Taller Límite Unidades")
        categoria = CategoriaResiduos.objects.create(
            nombre="Inorgánico límite manual",
            descripcion="Materiales por unidad",
            tipo_operacional=TIPO_INORGANICO,
        )
        subtipo = TipoResiduos.objects.create(
            categoria=categoria,
            nombre_residuo="Inorgánico límite manual",
            descripcion="Conteo",
        )
        Residuo.objects.create(
            tipo=categoria,
            subtipo=subtipo,
            unidad=995,
            contenedor_id="Unidades",
            estado=ESTADO_RESIDUO_CONFIRMADO,
            confirmado_at=timezone.now(),
            seccion="A1",
            profesor=self.user.username,
            asignatura="Taller previo",
            horario="08:00 - 09:00",
            numero_clase=1,
            created_by=self.user,
        )

        response = self.client.post(reverse("ingreso_manual_olvido"), {
            "fecha": clase.fecha.isoformat(),
            "taller": clase.id,
            "categoria": categoria.id_categoria,
            "subtipo": subtipo.id_tipo,
            "tipo_medicion": "unidades",
            "cantidad": "6",
            "motivo": "Intento sobre capacidad",
        })

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "solo admite 5 adicionales")
        self.assertFalse(
            Residuo.objects.filter(manual_reason="Intento sobre capacidad").exists()
        )


class SecurityControlsTests(TestCase):
    def setUp(self):
        cache.clear()
        self.admin_group, _ = Group.objects.get_or_create(name=ROL_ADMINISTRADOR)

    def tearDown(self):
        cache.clear()

    @override_settings(
        ENFORCE_INSTITUTIONAL_EMAIL_DOMAIN=True,
        INSTITUTIONAL_EMAIL_DOMAINS=["inacap.cl", "inacapmail.cl"],
    )
    def test_formulario_usuarios_rechaza_correo_no_institucional(self):
        form = UsuariosForm(data={
            "nombre": "Usuario",
            "apellido": "Externo",
            "email": "usuario@example.com",
            "grupo": self.admin_group.pk,
            "password1": "ClaveSegura123!",
            "password2": "ClaveSegura123!",
        })

        self.assertFalse(form.is_valid())
        self.assertIn("correo institucional", str(form.errors))

    def test_clave_de_usuario_exige_entre_8_y_12_caracteres(self):
        grupo, _ = Group.objects.get_or_create(name=ROL_PROFESOR)
        datos_base = {
            "nombre": "Usuario",
            "apellido": "Seguro",
            "email": "usuario.seguro@inacap.cl",
            "grupo": grupo.pk,
        }

        for clave in ("Aa1!bcDe", "Aa1!bcDef2$G"):
            with self.subTest(clave=clave):
                form = UsuariosForm(data={
                    **datos_base,
                    "password1": clave,
                    "password2": clave,
                })
                self.assertTrue(form.is_valid(), form.errors)

        for clave in ("Aa1!bcd", "Aa1!bcDef2$Gh"):
            with self.subTest(clave=clave):
                form = UsuariosForm(data={
                    **datos_base,
                    "password1": clave,
                    "password2": clave,
                })
                self.assertFalse(form.is_valid())
                self.assertIn("password1", form.errors)

    @override_settings(
        EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend",
        DEFAULT_FROM_EMAIL="sistema@inacap.cl",
    )
    def test_altas_de_todos_los_roles_persisten_acceso_perfil_y_correo(self):
        admin = User.objects.create_user(
            username="admin.pruebas@inacap.cl",
            email="admin.pruebas@inacap.cl",
            password="ClaveSegura123!",
        )
        admin.groups.add(self.admin_group)
        self.client.force_login(admin)

        casos = (
            (ROL_ESTUDIANTE, "alumno.nuevo@inacapmail.cl"),
            (ROL_PROFESOR, "profesor.nuevo@inacap.cl"),
            (ROL_ADMINISTRADOR, "admin.nuevo@inacap.cl"),
        )
        mail.outbox.clear()
        for indice, (rol, email) in enumerate(casos, start=1):
            grupo, _ = Group.objects.get_or_create(name=rol)
            response = self.client.post(reverse("usuarios"), {
                "nombre": f"Nombre{indice}",
                "apellido": f"Apellido{indice}",
                "email": email,
                "grupo": grupo.pk,
                "password1": "Aa1!bcD2",
                "password2": "Aa1!bcD2",
            })

            self.assertRedirects(response, reverse("usuarios"))
            perfil = Usuarios.objects.select_related("user").get(email=email)
            self.assertEqual(perfil.user.username, email)
            self.assertEqual(perfil.user.email, email)
            self.assertTrue(perfil.user.check_password("Aa1!bcD2"))
            self.assertEqual(
                list(perfil.user.groups.values_list("name", flat=True)),
                [rol],
            )
            mensaje = mail.outbox[-1]
            self.assertEqual(mensaje.to, [email])
            self.assertIn("ya está registrado", mensaje.body)
            self.assertNotIn("Aa1!bcD2", mensaje.body)

        self.assertEqual(len(mail.outbox), len(casos))

    @patch(
        "app.views.enviar_confirmacion_registro_usuario",
        side_effect=RuntimeError("SMTP no disponible"),
    )
    def test_falla_de_correo_no_revierte_el_usuario(self, _enviar):
        admin = User.objects.create_user(
            username="admin.pruebas@inacap.cl",
            email="admin.pruebas@inacap.cl",
            password="ClaveSegura123!",
        )
        admin.groups.add(self.admin_group)
        self.client.force_login(admin)
        grupo, _ = Group.objects.get_or_create(name=ROL_PROFESOR)

        response = self.client.post(
            reverse("usuarios"),
            {
                "nombre": "Docente",
                "apellido": "Confirmado",
                "email": "docente.confirmado@inacap.cl",
                "grupo": grupo.pk,
                "password1": "Aa1!bcD2",
                "password2": "Aa1!bcD2",
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(
            Usuarios.objects.filter(email="docente.confirmado@inacap.cl").exists()
        )
        self.assertTrue(
            User.objects.filter(username="docente.confirmado@inacap.cl").exists()
        )
        self.assertContains(
            response,
            "Usuario creado correctamente, pero no se pudo enviar el correo de confirmación.",
        )

    def test_usuarios_carga_popups_y_confirmacion_de_eliminacion(self):
        admin = User.objects.create_user(
            username="admin.interfaz@inacap.cl",
            email="admin.interfaz@inacap.cl",
            password="ClaveSegura123!",
        )
        admin.groups.add(self.admin_group)
        self.client.force_login(admin)
        Usuarios.objects.create(
            user=admin,
            nombre="Admin",
            apellido="Interfaz",
            email=admin.email,
        )

        response = self.client.get(reverse("usuarios"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "js/app_popups.js")
        self.assertContains(response, 'data-confirm-title="Eliminar usuario"')
        self.assertContains(response, "data-confirm-message=")

    @override_settings(HEALTH_CHECK_REQUIRE_LOGIN=True, HEALTH_CHECK_EXPOSE_DETAILS=False)
    def test_health_check_requiere_login_si_esta_configurado(self):
        response = self.client.get(reverse("health_check"))
        self.assertEqual(response.status_code, 403)

        user = User.objects.create_user(
            username="admin@test.cl",
            email="admin@test.cl",
            password="ClaveSegura123!",
        )
        user.groups.add(self.admin_group)
        self.client.force_login(user)

        response = self.client.get(reverse("health_check"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["ok"], True)

    @override_settings(PASSWORD_MAX_AGE_DAYS=120)
    def test_login_rechaza_contrasena_vencida_a_los_120_dias(self):
        user = User.objects.create_user(
            username="docente@inacap.cl",
            email="docente@inacap.cl",
            password="ClaveSegura123!",
        )
        User.objects.filter(pk=user.pk).update(
            date_joined=timezone.now() - timedelta(days=121)
        )

        response = self.client.post(reverse("login"), {
            "username": user.username,
            "password": "ClaveSegura123!",
        })

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "La contraseña venció")
        self.assertNotIn("_auth_user_id", self.client.session)
        self.assertTrue(
            AuditLog.objects.filter(
                action="login_password_expired",
                model_name="User",
                object_id=str(user.pk),
            ).exists()
        )

    @override_settings(LOGIN_MAX_ATTEMPTS=3, LOGIN_LOCKOUT_SECONDS=60)
    def test_login_bloquea_temporalmente_intentos_repetidos(self):
        user = User.objects.create_user(
            username="bloqueo@inacap.cl",
            email="bloqueo@inacap.cl",
            password="ClaveSegura123!",
        )

        for _ in range(2):
            response = self.client.post(reverse("login"), {
                "username": user.username,
                "password": "incorrecta",
            })
            self.assertEqual(response.status_code, 200)

        response = self.client.post(reverse("login"), {
            "username": user.username,
            "password": "incorrecta",
        })
        self.assertEqual(response.status_code, 429)
        self.assertContains(response, "Demasiados intentos fallidos", status_code=429)

        response = self.client.post(reverse("login"), {
            "username": user.username,
            "password": "ClaveSegura123!",
        })
        self.assertEqual(response.status_code, 429)
        self.assertTrue(AuditLog.objects.filter(action="login_rate_limited").exists())

    def test_respuestas_incluyen_cabeceras_de_seguridad(self):
        response = self.client.get(reverse("login"))

        self.assertEqual(response.headers["X-Content-Type-Options"], "nosniff")
        self.assertEqual(response.headers["X-Frame-Options"], "DENY")
        self.assertEqual(response.headers["Referrer-Policy"], "same-origin")

    @override_settings(TRUSTED_PROXY_IPS=set())
    def test_ip_reenviada_se_ignora_si_el_proxy_no_es_confiable(self):
        self.client.post(
            reverse("login"),
            {"username": "nadie", "password": "incorrecta"},
            REMOTE_ADDR="198.51.100.10",
            HTTP_X_FORWARDED_FOR="203.0.113.25",
        )
        auditoria = AuditLog.objects.filter(action="login_failed").latest("created_at")
        self.assertEqual(str(auditoria.ip_address), "198.51.100.10")

    @override_settings(TRUSTED_PROXY_IPS={"127.0.0.1"})
    def test_ip_reenviada_se_acepta_desde_proxy_confiable(self):
        self.client.post(
            reverse("login"),
            {"username": "nadie", "password": "incorrecta"},
            REMOTE_ADDR="127.0.0.1",
            HTTP_X_FORWARDED_FOR="203.0.113.25",
        )
        auditoria = AuditLog.objects.filter(action="login_failed").latest("created_at")
        self.assertEqual(str(auditoria.ip_address), "203.0.113.25")


class BackupSecurityTests(SimpleTestCase):
    def _create_database(self, path):
        with closing(sqlite3.connect(path)) as connection:
            connection.execute("CREATE TABLE registros (valor TEXT NOT NULL)")
            connection.execute("INSERT INTO registros (valor) VALUES ('original')")
            connection.commit()

    def test_base_sqlite_en_memoria_no_intenta_respaldo_fisico(self):
        database_settings = {
            "default": {
                "ENGINE": "django.db.backends.sqlite3",
                "NAME": "file:memorydb_default?mode=memory&cache=shared",
            },
        }
        with self.settings(DATABASES=database_settings):
            self.assertTrue(is_sqlite_memory_database())

    def test_respaldo_verificado_detecta_alteracion_y_restaura(self):
        with TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir)
            database = root / "db.sqlite3"
            backup_dir = root / "backups"
            self._create_database(database)
            database_settings = {
                "default": {
                    "ENGINE": "django.db.backends.sqlite3",
                    "NAME": database,
                }
            }

            with self.settings(DATABASES=database_settings, BACKUP_DIR=backup_dir):
                backup = create_verified_backup()
                self.assertTrue(backup.exists())
                self.assertTrue(backup.with_suffix(".sqlite3.sha256").exists())
                self.assertTrue(verify_sqlite_backup(backup, require_checksum=True))

                with closing(sqlite3.connect(database)) as connection:
                    connection.execute("INSERT INTO registros (valor) VALUES ('posterior')")
                    connection.commit()

                restored, pre_restore = restore_verified_backup(backup)
                self.assertEqual(restored, database.resolve())
                self.assertIsNotNone(pre_restore)
                with closing(sqlite3.connect(database)) as connection:
                    values = connection.execute(
                        "SELECT valor FROM registros ORDER BY rowid"
                    ).fetchall()
                self.assertEqual(values, [("original",)])

                checksum_path = backup.with_suffix(".sqlite3.sha256")
                checksum_path.write_text("0" * 64, encoding="ascii")
                with self.assertRaisesMessage(ValueError, "SHA-256"):
                    verify_sqlite_backup(backup, require_checksum=True)

    def test_respaldo_unico_se_sobrescribe_sin_crear_versiones(self):
        with TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir)
            database = root / "db.sqlite3"
            backup_dir = root / "backups"
            self._create_database(database)
            database_settings = {
                "default": {
                    "ENGINE": "django.db.backends.sqlite3",
                    "NAME": database,
                },
            }

            with self.settings(
                DATABASES=database_settings,
                BACKUP_DIR=backup_dir,
                ROLLING_BACKUP_FILENAME="respaldo_base_datos.sqlite3",
            ):
                primero = create_rolling_backup()
                with closing(sqlite3.connect(database)) as connection:
                    connection.execute(
                        "INSERT INTO registros (valor) VALUES ('actualizado')"
                    )
                    connection.commit()
                segundo = create_rolling_backup()

            self.assertEqual(primero, segundo)
            self.assertEqual(
                [path.name for path in backup_dir.glob("*.sqlite3")],
                ["respaldo_base_datos.sqlite3"],
            )
            self.assertTrue(verify_sqlite_backup(segundo, require_checksum=True))
            with closing(sqlite3.connect(segundo)) as connection:
                valores = connection.execute(
                    "SELECT valor FROM registros ORDER BY rowid"
                ).fetchall()
            self.assertEqual(valores, [("original",), ("actualizado",)])


class ResetOperationalDataCommandTests(TransactionTestCase):
    reset_sequences = True

    def setUp(self):
        self.group, _ = Group.objects.get_or_create(name="Administrador")
        self.user = User.objects.create_user(
            username="conservar@test.cl",
            email="conservar@test.cl",
            password="ClaveSegura123!",
        )
        self.user.groups.add(self.group)
        Usuarios.objects.create(
            user=self.user,
            nombre="Usuario",
            apellido="Conservado",
            email=self.user.email,
        )
        self.categoria = CategoriaResiduos.objects.create(
            nombre="Orgánicos",
            descripcion="Configuración maestra",
            tipo_operacional=TIPO_ORGANICO,
        )
        self.subtipo = TipoResiduos.objects.create(
            nombre_residuo="Orgánico de prueba",
            descripcion="Configuración maestra",
            categoria=self.categoria,
        )
        self.destino = Destino.objects.create(
            nombre="Compostera",
            categoria=self.categoria,
        )
        self.taller = ClaseHorario.objects.create(
            seccion="TEST-01",
            profesor="Profesor",
            asignatura="Taller de prueba",
            horario="08:00 - 09:00",
            fecha=timezone.localdate(),
            dia_semana=timezone.localdate().weekday(),
            hora_inicio=time(8, 0),
            hora_fin=time(9, 0),
        )
        self.residuo = Residuo.objects.create(
            taller=self.taller,
            seccion=self.taller.seccion,
            profesor=self.taller.profesor,
            asignatura=self.taller.asignatura,
            horario=self.taller.horario,
            numero_clase=1,
            tipo=self.categoria,
            subtipo=self.subtipo,
            peso=Decimal("1.250"),
        )
        HistorialRetiro.objects.create(
            usuario=self.user,
            contenedor_origen="organico",
            tipo_residuo="Orgánico",
            destino=self.destino,
            cantidad_peso=Decimal("1.25"),
        )
        WeightReading.objects.create(
            weight_kg=Decimal("1.250"),
            created_by=self.user,
        )
        Actividad.objects.create(
            profesor="Profesor",
            fecha=timezone.localdate(),
            nombre="Actividad histórica",
        )
        AuditLog.objects.create(
            user=self.user,
            action="dato_operativo_de_prueba",
        )
        OperationalLock.objects.update_or_create(
            name="organico",
            defaults={"version": 8},
        )

    @patch(
        "app.management.commands.reset_operational_data.create_rolling_backup",
        return_value=Path("respaldo_base_datos.sqlite3"),
    )
    @patch(
        "app.management.commands.reset_operational_data.create_verified_backup",
        return_value=Path("antes_limpieza.sqlite3"),
    )
    def test_limpieza_conserva_usuarios_y_configuracion_maestra(
        self,
        backup_previo,
        backup_unico,
    ):
        salida = StringIO()
        call_command("reset_operational_data", "--confirm", stdout=salida)

        self.assertTrue(User.objects.filter(pk=self.user.pk).exists())
        self.assertTrue(Usuarios.objects.filter(user=self.user).exists())
        self.assertTrue(Group.objects.filter(pk=self.group.pk).exists())
        self.assertTrue(CategoriaResiduos.objects.filter(pk=self.categoria.pk).exists())
        self.assertTrue(TipoResiduos.objects.filter(pk=self.subtipo.pk).exists())
        self.assertTrue(Destino.objects.filter(pk=self.destino.pk).exists())

        self.assertFalse(ClaseHorario.objects.exists())
        self.assertFalse(Residuo.objects.exists())
        self.assertFalse(HistorialRetiro.objects.exists())
        self.assertFalse(WeightReading.objects.exists())
        self.assertFalse(Actividad.objects.exists())
        self.assertFalse(AuditLog.objects.exists())
        self.assertEqual(
            set(OperationalLock.objects.values_list("name", flat=True)),
            {"organico", "unidades"},
        )
        self.assertFalse(
            OperationalLock.objects.exclude(version=0).exists()
        )
        backup_previo.assert_called_once()
        backup_unico.assert_called_once()
        self.assertIn("Puesta a cero completada", salida.getvalue())

    @patch("app.management.commands.reset_operational_data.create_verified_backup")
    def test_sin_confirmacion_solo_simula(self, backup_previo):
        salida = StringIO()
        call_command("reset_operational_data", stdout=salida)

        self.assertTrue(Residuo.objects.filter(pk=self.residuo.pk).exists())
        backup_previo.assert_not_called()
        self.assertIn("Simulacion", salida.getvalue())


class GestionCategoriasTests(TestCase):
    def setUp(self):
        grupo, _ = Group.objects.get_or_create(name=ROL_ADMINISTRADOR)
        self.user = User.objects.create_user(
            username="admin-categorias@test.cl",
            email="admin-categorias@test.cl",
            password="ClaveSegura123!",
        )
        self.user.groups.add(grupo)
        self.client.force_login(self.user)
        self.categoria = CategoriaResiduos.objects.filter(
            tipo_operacional=TIPO_ORGANICO,
        ).order_by("id_categoria").first()
        if self.categoria is None:
            self.categoria = CategoriaResiduos.objects.create(
                nombre="Orgánico",
                descripcion="Residuos compostables",
                tipo_operacional=TIPO_ORGANICO,
            )
        self.subcategoria = TipoResiduos.objects.get(nombre_residuo="Orgánico")

    def test_listado_categorias_no_muestra_gestion_de_subcategorias(self):
        response = self.client.get(reverse("gestionCategorias"))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Subcategoría")
        self.assertNotContains(response, "Subcategorías")
        self.assertNotContains(response, 'id="campos-subcategoria"')
        self.assertNotContains(response, 'value="crear_subcategoria"')

    def test_post_de_subcategoria_ya_no_crea_registros(self):
        response = self.client.post(reverse("gestionCategorias"), {
            "accion": "crear_subcategoria",
            "subcategoria-categoria": self.categoria.pk,
            "subcategoria-nombre_residuo": "Verduras prueba",
            "subcategoria-descripcion": "Verduras descartadas",
        })

        self.assertEqual(response.status_code, 200)
        self.assertFalse(
            TipoResiduos.objects.filter(
                categoria=self.categoria,
                nombre_residuo="Verduras prueba",
            ).exists()
        )

    def test_subcategoria_historica_se_conserva_sin_mostrar_gestion(self):
        self.subcategoria = TipoResiduos.objects.create(
            categoria=self.categoria,
            nombre_residuo="Frutas prueba",
            descripcion="Frutas descartadas",
        )
        residuo = Residuo.objects.create(
            seccion="A1",
            profesor="Profesor Demo",
            asignatura="Taller Compost",
            horario="10:00 - 11:00",
            numero_clase=1,
            tipo=self.categoria,
            subtipo=self.subcategoria,
            peso=Decimal("1.000"),
            estado=ESTADO_RESIDUO_CONFIRMADO,
        )

        response = self.client.get(reverse("gestionCategorias"))

        self.assertEqual(response.status_code, 200)
        self.assertTrue(TipoResiduos.objects.filter(pk=self.subcategoria.pk).exists())
        residuo.refresh_from_db()
        self.assertEqual(residuo.tipo, self.categoria)
        self.assertEqual(residuo.subtipo, self.subcategoria)
        self.assertNotContains(response, self.subcategoria.nombre_residuo)
