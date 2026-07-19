from django.conf import settings
from django.core.cache import cache
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.utils.text import slugify
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET, require_POST
from datetime import datetime, time, timedelta
from pathlib import Path
from .models import (
    ESTADO_RESIDUO_ANULADO,
    ESTADO_RESIDUO_CONFIRMADO,
    ESTADO_RESIDUO_PENDIENTE,
    AuditLog,
    CategoriaResiduos,
    ClaseHorario,
    Destino,
    HistorialRetiro,
    OperationalLock,
    Residuo,
    TipoResiduos,
    Usuarios,
    WeightReading,
)
from pytz import timezone as tz
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.contrib.auth.models import User
from .forms import (
    ClaseHorarioForm,
    DestinoForm,
    MEDICION_UNIDADES,
    ManualResiduoOlvidoForm,
    UsuariosForm,
    CategoriaResiduosForm,
)
from .roles import (
    aplicar_privilegios_por_rol,
    es_administrador,
    es_operador,
    es_profesor,
    puede_crear_talleres,
    puede_ingresar_manual,
    puede_gestionar_horarios,
)
from .residue_types import (
    TIPO_INORGANICO,
    TIPO_ORGANICO,
    asegurar_tipos_residuos_predeterminados,
    normalizar_tipo_operacional,
    tipo_operacional_categoria,
)
from .services import balanza_service
from .services.backup_service import create_rolling_backup, is_sqlite_memory_database
from .services.email_service import enviar_confirmacion_registro_usuario
from .services.dashboard import (
    LIMITE_UNIDADES_GLOBAL,
    crear_dashboard_context,
    distribuir_peso_organico,
    obtener_datos_unidades,
)
from django.db import transaction, IntegrityError, OperationalError
from django.db.models import F, Max, Q, Sum
import json, re, hashlib, logging, qrcode, secrets, zipfile
from decimal import Decimal, InvalidOperation
from urllib.parse import urljoin
from html import escape as xml_escape
from django.utils import timezone
from PIL import Image
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from .serializers import DatosClaseSerializer
from rest_framework.decorators import api_view
from rest_framework.response import Response

logger = logging.getLogger(__name__)

admin_required = user_passes_test(es_administrador, login_url="login")
manual_entry_required = user_passes_test(puede_ingresar_manual, login_url="login")

PROFESOR_TALLERES_PRECARGADOS = "smartadmin"
TALLERES_PRECARGADOS = (
    ("Taller 1", time(8, 0), time(12, 40)),
    ("Taller 2", time(13, 40), time(18, 20)),
    ("Taller 3", time(18, 30), time(22, 30)),
)


def _client_ip(request):
    if request is None:
        return None
    remote_addr = (request.META.get("REMOTE_ADDR") or "").strip() or None
    trusted_proxies = getattr(settings, "TRUSTED_PROXY_IPS", set())
    forwarded_for = request.META.get("HTTP_X_FORWARDED_FOR")
    if remote_addr in trusted_proxies and forwarded_for:
        client_ip = forwarded_for.split(",")[0].strip()
        if client_ip:
            return client_ip
    return remote_addr


def _clean_text(value, max_length, required=True):
    value = (value or "").strip()
    if required and not value:
        raise ValueError("Campo obligatorio incompleto.")
    return value[:max_length]


def _parse_decimal(value, field_name, min_value=None, max_value=None, required=True):
    if value in (None, "", "null"):
        if required:
            raise ValueError(f"{field_name} es obligatorio.")
        return None

    try:
        parsed = Decimal(str(value).replace(",", "."))
    except (InvalidOperation, ValueError):
        raise ValueError(f"{field_name} debe ser un número válido.")

    if min_value is not None and parsed < Decimal(str(min_value)):
        raise ValueError(f"{field_name} no puede ser menor que {min_value}.")
    if max_value is not None and parsed > Decimal(str(max_value)):
        raise ValueError(f"{field_name} no puede ser mayor que {max_value}.")

    return parsed.quantize(Decimal("0.001"))


def _parse_int(value, field_name, min_value=None, max_value=None, required=True):
    if value in (None, "", "null"):
        if required:
            raise ValueError(f"{field_name} es obligatorio.")
        return None

    try:
        parsed = int(value)
    except (TypeError, ValueError):
        raise ValueError(f"{field_name} debe ser un número entero.")

    if min_value is not None and parsed < min_value:
        raise ValueError(f"{field_name} no puede ser menor que {min_value}.")
    if max_value is not None and parsed > max_value:
        raise ValueError(f"{field_name} no puede ser mayor que {max_value}.")

    return parsed


def _categorias_residuo_disponibles():
    ids = [
        CategoriaResiduos.objects.filter(tipo_operacional=tipo)
        .order_by("-id_categoria")
        .values_list("id_categoria", flat=True)
        .first()
        for tipo in (TIPO_ORGANICO, TIPO_INORGANICO)
    ]
    return CategoriaResiduos.objects.filter(
        id_categoria__in=[categoria_id for categoria_id in ids if categoria_id]
    ).order_by("id_categoria")


def _subtipo_interno_categoria(categoria):
    """Mantiene compatibilidad histórica sin pedir un subtipo al usuario."""
    if categoria is None:
        return None

    subtipos = TipoResiduos.objects.filter(categoria=categoria).order_by("id_tipo")
    return (
        subtipos.filter(nombre_residuo__iexact=categoria.nombre).first()
        or subtipos.first()
    )


def _audit(request, action, instance=None, metadata=None):
    user = getattr(request, "user", None)
    if user is not None and not user.is_authenticated:
        user = None

    AuditLog.objects.create(
        user=user,
        action=action,
        model_name=instance.__class__.__name__ if instance is not None else "",
        object_id=str(getattr(instance, "pk", "")) if instance is not None else "",
        ip_address=_client_ip(request),
        user_agent=((request.META.get("HTTP_USER_AGENT") or "")[:255] if request else ""),
        metadata=metadata or {},
    )


def _safe_next_url(request):
    next_url = request.POST.get("next") or request.GET.get("next")
    if next_url and url_has_allowed_host_and_scheme(
        url=next_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return next_url
    return None


def _password_expired(user):
    max_age_days = getattr(settings, "PASSWORD_MAX_AGE_DAYS", 0)
    if not max_age_days:
        return False

    perfil = getattr(user, "perfil_usuario", None)
    changed_at = getattr(perfil, "password_changed_at", None) or user.date_joined
    return changed_at <= timezone.now() - timedelta(days=max_age_days)


def _login_throttle_key(request, username):
    identity = f"{_client_ip(request) or 'sin-ip'}|{(username or '').strip().lower()}"
    return hashlib.sha256(identity.encode("utf-8")).hexdigest()


def _login_is_locked(throttle_key):
    return bool(cache.get(f"login-lock:{throttle_key}"))


def _record_login_failure(throttle_key):
    attempts_key = f"login-attempts:{throttle_key}"
    timeout = max(int(getattr(settings, "LOGIN_LOCKOUT_SECONDS", 900)), 60)
    try:
        attempts = cache.incr(attempts_key)
    except ValueError:
        cache.set(attempts_key, 1, timeout=timeout)
        attempts = 1

    if attempts >= max(int(getattr(settings, "LOGIN_MAX_ATTEMPTS", 5)), 1):
        cache.set(f"login-lock:{throttle_key}", True, timeout=timeout)
        cache.delete(attempts_key)
        return True
    return False


def _clear_login_failures(throttle_key):
    cache.delete_many([
        f"login-attempts:{throttle_key}",
        f"login-lock:{throttle_key}",
    ])


def _weight_max():
    try:
        return Decimal(str(settings.MAX_WEIGHT_KG))
    except (InvalidOperation, TypeError):
        return Decimal("1000")


def _bloquear_capacidad(nombre):
    """Serializa cambios de capacidad en SQLite y bases con bloqueo por fila."""
    try:
        bloqueo = OperationalLock.objects.select_for_update().get(name=nombre)
    except OperationalLock.DoesNotExist:
        bloqueo = OperationalLock.objects.create(name=nombre)
    OperationalLock.objects.filter(pk=bloqueo.pk).update(version=F("version") + 1)


def _validar_capacidad_unidades(unidades_nuevas, residuo_excluido_id=None):
    residuos = Residuo.objects.filter(
        estado=ESTADO_RESIDUO_CONFIRMADO,
        retirado=False,
        tipo__tipo_operacional=TIPO_INORGANICO,
    )
    if residuo_excluido_id:
        residuos = residuos.exclude(pk=residuo_excluido_id)

    unidades_actuales = int(residuos.aggregate(total=Sum("unidad"))["total"] or 0)
    unidades_nuevas = int(unidades_nuevas or 0)
    total_resultante = unidades_actuales + unidades_nuevas
    if total_resultante > LIMITE_UNIDADES_GLOBAL:
        disponibles = max(LIMITE_UNIDADES_GLOBAL - unidades_actuales, 0)
        raise ValueError(
            f"El almacén tiene {unidades_actuales} unidades y solo admite "
            f"{disponibles} adicionales. El límite total es "
            f"{LIMITE_UNIDADES_GLOBAL}."
        )
    return total_resultante


def _puede_gestionar_pendiente(user, residuo):
    return es_administrador(user) or (
        residuo.created_by_id is not None and residuo.created_by_id == user.id
    )


def _crear_lectura_peso(request, data, user=None):
    weight = _parse_decimal(data.get("weight_kg"), "weight_kg", min_value=0, max_value=_weight_max())
    device_name = _clean_text(data.get("device_name") or "sin-dispositivo", 100, required=False)
    raw = _clean_text(data.get("raw_data") or "", 1000, required=False)
    is_stable_raw = data.get("is_stable", True)
    if isinstance(is_stable_raw, str):
        is_stable = is_stable_raw.strip().lower() not in {"0", "false", "no", "off"}
    else:
        is_stable = bool(is_stable_raw)

    reading = WeightReading.objects.create(
        weight_kg=weight,
        device_name=device_name,
        raw_data=raw,
        is_stable=is_stable,
        created_by=user,
        source_ip=_client_ip(request),
    )
    _audit(request, "weight_reading_created", reading, {"device_name": device_name, "is_stable": is_stable})
    return reading


def _ultima_lectura_peso_reciente():
    try:
        max_age_seconds = int(getattr(settings, "WEIGHT_READING_MAX_AGE_SECONDS", 10))
    except (TypeError, ValueError):
        max_age_seconds = 10
    if max_age_seconds <= 0:
        return None
    return (
        WeightReading.objects
        .filter(
            is_stable=True,
            created_at__gte=timezone.now() - timedelta(seconds=max_age_seconds),
        )
        .order_by("-created_at")
        .first()
    )


def _lectura_peso_json(reading, fuente="bridge"):
    return {
        "ok": True,
        "peso": float(reading.weight_kg),
        "unidad": "kg",
        "puerto": None,
        "baudrate": None,
        "serial_mode": None,
        "line_control": None,
        "mensaje": "Lectura estable recibida automáticamente.",
        "fuente": fuente,
        "dispositivo": reading.device_name,
        "created_at": timezone.localtime(reading.created_at).isoformat(),
    }


def _error_balanza_json(resultado=None):
    resultado = resultado or {}
    return {
        "ok": False,
        "peso": None,
        "unidad": None,
        "puerto": resultado.get("puerto"),
        "baudrate": resultado.get("baudrate"),
        "serial_mode": resultado.get("serial_mode"),
        "line_control": resultado.get("line_control"),
        "codigo": resultado.get("codigo", "error_balanza"),
        "mensaje": resultado.get("mensaje") or "No se pudo leer la balanza.",
        "accion": resultado.get("accion") or (
            "Revise conexión, controlador USB-serial, alimentación y que el puerto no esté ocupado."
        ),
        "puertos_probados": resultado.get("puertos_probados", []),
    }


def _public_absolute_uri(request, path):
    public_base_url = getattr(settings, "PUBLIC_BASE_URL", "").strip()
    if public_base_url:
        return urljoin(public_base_url.rstrip("/") + "/", path.lstrip("/"))

    return request.build_absolute_uri(path)


def _buscar_usuario_login(identificador):
    identificador = (identificador or "").strip()
    if not identificador:
        return None

    user = User.objects.filter(username__iexact=identificador).first()
    if user:
        return user

    return User.objects.filter(email__iexact=identificador, is_active=True).first()


def _sincronizar_auth_user(usuario, password=None, grupo=None):
    user = usuario.user

    if not user:
        user = User.objects.filter(username__iexact=usuario.email).first()
    if not user:
        user = User.objects.create_user(username=usuario.email, email=usuario.email)

    user.username = usuario.email
    user.email = usuario.email
    user.first_name = usuario.nombre
    user.last_name = usuario.apellido
    user.is_active = True

    profile_update_fields = []
    if password:
        user.set_password(password)
        usuario.password_changed_at = timezone.now()
        profile_update_fields.append("password_changed_at")

    aplicar_privilegios_por_rol(user, grupo)
    user.save()

    if grupo:
        user.groups.set([grupo])

    if usuario.user_id != user.id:
        usuario.user = user
        profile_update_fields.append("user")

    if profile_update_fields:
        usuario.save(update_fields=profile_update_fields)

    return user


def _datos_clase_actual():
    ahora = datetime.now(tz('America/Santiago'))
    hora_actual_str = ahora.strftime("%d-%m-%Y %H:%M:%S")

    dia_semana = ahora.weekday()
    hora = ahora.time()

    hoy = ahora.date()

    # --- Obtener las clases activas del día ---
    clases_del_dia = ClaseHorario.objects.filter(
        archivado=False,
        dia_semana=dia_semana,
    ).filter(Q(fecha=hoy) | Q(fecha__isnull=True)).exclude(
        profesor=PROFESOR_TALLERES_PRECARGADOS
    )

    clases_en_horario = []

    for c in clases_del_dia:
        inicio = c.hora_inicio
        fin = c.hora_fin

        # Caso normal
        if inicio <= fin:
            if inicio <= hora <= fin:
                clases_en_horario.append(c)

        # Caso cruzando medianoche (ej: 20:00 → 00:00)
        else:
            if hora >= inicio or hora <= fin:
                clases_en_horario.append(c)

    # --- Sin clases ---
    if not clases_en_horario:
        return {"fuera_horario": True, "multiple": False}, True

    # --- Varias clases ---
    if len(clases_en_horario) > 1:
        lista = [
            {
                "id": c.id,
                "seccion": c.seccion,
                "profesor": c.profesor,
                "asignatura": c.asignatura,
                "horario": c.horario,
            }
            for c in clases_en_horario
        ]

        return {
            "multiple": True,
            "actividades": lista,
            "fuera_horario": False,
        }, False

    # --- Una sola clase ---
    clase = clases_en_horario[0]

    numero_actual = Residuo.objects.filter(
        asignatura=clase.asignatura,
        seccion=clase.seccion
    ).count()

    return {
        "id": clase.id,
        "multiple": False,
        "seccion": clase.seccion,
        "profesor": clase.profesor,
        "asignatura": clase.asignatura,
        "horario": clase.horario,
        "numero_clase": numero_actual + 1,
        "hora_escaneo": hora_actual_str,
        "fuera_horario": False,
    }, False


def _clase_disponible_para_registro(clase, ahora=None):
    ahora = ahora or datetime.now(tz("America/Santiago"))
    return bool(
        not clase.archivado
        and clase.profesor != PROFESOR_TALLERES_PRECARGADOS
        and clase.fecha == ahora.date()
        and clase.dia_semana == ahora.weekday()
        and _hora_en_rango(ahora.time(), clase.hora_inicio, clase.hora_fin)
    )



@login_required(login_url="login")
def obtener_datos(request):
    data, fuera_horario = _datos_clase_actual()

    # si NO hay clases → fuera del horario
    if fuera_horario:
        return JsonResponse({"fuera_horario": True})

    # si hay conflicto
    if data.get("multiple", False):
        return JsonResponse({
            "fuera_horario": False,
            "conflicto": True,
            "actividades": data["actividades"]
        })

    # si hay solo una clase
    return JsonResponse({
        "id": data["id"],
        "fuera_horario": False,
        "conflicto": False,
        "nombre": data["asignatura"],
        "descripcion": data["seccion"],
        "profesor": data["profesor"],
        "horario": data["horario"],
        "numero_clase": data["numero_clase"],
        "hora_escaneo": data["hora_escaneo"]
    })


@login_required(login_url="login")
@api_view(["GET"])
def obtener_datos_api(request):
    data, fuera = _datos_clase_actual()
    serializer = DatosClaseSerializer(data)
    return Response(serializer.data)

@login_required(login_url="login")
@api_view(["GET"])
def seleccionar_actividad(request, id):
    clase = get_object_or_404(ClaseHorario, id=id)
    if not _clase_disponible_para_registro(clase):
        return Response(
            {"error": "El taller seleccionado ya no está activo."},
            status=409,
        )

    numero_actual = Residuo.objects.filter(
        asignatura=clase.asignatura,
        seccion=clase.seccion
    ).count()

    numero_clase = numero_actual + 1

    ahora = datetime.now(tz('America/Santiago'))
    hora_actual_str = ahora.strftime("%d-%m-%Y %H:%M:%S")

    return Response({
        "id": clase.id,
        "asignatura": clase.asignatura,
        "seccion": clase.seccion,
        "profesor": clase.profesor,
        "horario": clase.horario,
        "numero_clase": numero_clase,
        "hora_escaneo": hora_actual_str,
    })

@login_required(login_url="login")
@require_POST
def guardar(request):
    if not es_operador(request.user):
        return JsonResponse({"success": False, "error": "No tiene permisos para registrar residuos."}, status=403)

    try:
        taller_id = _parse_int(request.POST.get("taller_id"), "taller_id", min_value=1)

        with transaction.atomic():
            clase = get_object_or_404(
                ClaseHorario.objects.select_for_update(),
                pk=taller_id,
            )
            if not _clase_disponible_para_registro(clase):
                raise ValueError("El taller seleccionado ya no está activo.")

            seccion = _clean_text(clase.seccion, 200)
            profesor = _clean_text(clase.profesor, 100)
            asignatura = _clean_text(clase.asignatura, 100)
            horario = _clean_text(clase.horario, 100)
            ultimo_numero = (
                Residuo.objects
                .filter(asignatura=asignatura, seccion=seccion)
                .aggregate(maximo=Max("numero_clase"))["maximo"] or 0
            )
            numero_clase = int(ultimo_numero) + 1
            if numero_clase > 9999:
                raise ValueError("Se alcanzó el máximo de registros para este taller.")

            residuo = Residuo.objects.create(
                taller=clase,
                seccion=seccion,
                profesor=profesor,
                asignatura=asignatura,
                horario=horario,
                numero_clase=numero_clase,
                estado=ESTADO_RESIDUO_PENDIENTE,
                hora_escaneo=timezone.now(),
                created_by=request.user,
                updated_by=request.user,
                source_ip=_client_ip(request),
            )
        _audit(request, "residuo_created", residuo)

        return JsonResponse({
            "success": True,
            "id": residuo.id,
            "numero_clase": residuo.numero_clase,
            "registro_contexto": {
                "taller_id": clase.id,
                "seccion": clase.seccion,
                "profesor": clase.profesor,
                "asignatura": clase.asignatura,
                "horario": clase.horario,
            },
            "redirect": reverse("residuos")
        })
    except IntegrityError:
        return JsonResponse({
            "success": False,
            "error": "Este taller ya fue registrado para esa clase."
        }, status=409)
    except ValueError as error:
        return JsonResponse({"success": False, "error": str(error)}, status=400)




@login_required(login_url="login")
def taller(request):
    if not es_operador(request.user):
        messages.error(request, "No tiene permisos para ingresar al formulario de registro.")
        return redirect("dashboard")

    return render(request, 'formulario.html')

def fuera_horario(request):
    return render(request, 'fuerahorario.html')


@login_required(login_url="login")
def residuos(request):
    if not es_operador(request.user):
        messages.error(request, "No tiene permisos para registrar residuos.")
        return redirect("dashboard")

    asegurar_tipos_residuos_predeterminados()
    categorias_residuo = list(_categorias_residuo_disponibles())
    tipos_disponibles = {
        categoria.tipo_operacional for categoria in categorias_residuo
    }
    if not {TIPO_ORGANICO, TIPO_INORGANICO}.issubset(tipos_disponibles):
        messages.error(
            request,
            "Deben estar configurados los tipos de residuo Orgánico e Inorgánico.",
        )
        return redirect("dashboard")

    pendiente_reanudado = None
    pendiente_id = request.GET.get("pendiente")
    if pendiente_id:
        try:
            pendiente_id = _parse_int(pendiente_id, "pendiente", min_value=1)
        except ValueError as error:
            messages.error(request, str(error))
            return redirect("dashboard")

        pendiente = get_object_or_404(
            Residuo.objects.select_related("taller", "tipo", "subtipo"),
            pk=pendiente_id,
            estado=ESTADO_RESIDUO_PENDIENTE,
        )
        if not _puede_gestionar_pendiente(request.user, pendiente):
            messages.error(request, "No puede continuar un registro pendiente creado por otro usuario.")
            return redirect("dashboard")

        pendiente_reanudado = {
            "residuo_id": pendiente.id,
            "tipo_id": pendiente.tipo_id,
            "registro_contexto": {
                "taller_id": pendiente.taller_id,
                "seccion": pendiente.seccion,
                "profesor": pendiente.profesor,
                "asignatura": pendiente.asignatura,
                "horario": pendiente.horario,
            },
        }

    return render(request, "residuo.html", {
        "categorias_residuo": categorias_residuo,
        "pendiente_reanudado": pendiente_reanudado,
    })

def ver_qr(request):
    if request.user.is_authenticated:
        return redirect("dashboard")
    return redirect("login")


@login_required(login_url="login")
def pagina_codigo_qr(request):
    registro_path = reverse("registrar_evento")
    registro_url = _public_absolute_uri(request, registro_path)

    return render(request, "codigo_qr.html", {
        "registro_path": registro_path,
        "registro_url": registro_url,
        "conexion_segura": request.is_secure() or registro_url.startswith("https://"),
    })


@login_required(login_url="login")
def acceso_profesor_qr(request):
    if not es_profesor(request.user):
        messages.error(request, "El acceso por QR de profesor está permitido solo para usuarios con rol Profesor.")
        return redirect("dashboard" if request.user.is_authenticated else "login")

    return redirect("profesor")


def generar_qr_profesor(request):
    url = _public_absolute_uri(request, reverse("acceso_profesor_qr"))
    qr = qrcode.QRCode(version=1, box_size=8, border=4)
    qr.add_data(url)
    qr.make(fit=True)

    img = qr.make_image(fill="black", back_color="white")
    response = HttpResponse(content_type="image/png")
    img.save(response, "PNG")
    return response


@login_required(login_url="login")
def generar_codigo(request):
    ruta = request.GET.get("ruta") or reverse("registrar_evento")
    if not ruta.startswith("/") or ruta.startswith("//"):
        ruta = reverse("registrar_evento")

    url = _public_absolute_uri(request, ruta)
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(url)
    qr.make(fit=True)

    img = qr.make_image(fill="black", back_color="white")
    response = HttpResponse(content_type="image/png")
    img.save(response, "PNG")
    return response


def login_view(request):
    if request.user.is_authenticated:
        return redirect("dashboard")

    if request.method == "POST":
        username = (request.POST.get("username") or "").strip()
        password = request.POST.get("password")
        throttle_key = _login_throttle_key(request, username)

        if _login_is_locked(throttle_key):
            _audit(request, "login_rate_limited", metadata={"username": username})
            messages.error(
                request,
                "Demasiados intentos fallidos. Espere 15 minutos antes de volver a intentar."
            )
            return render(request, "login.html", {
                "next": request.GET.get("next", ""),
            }, status=429)

        user_obj = _buscar_usuario_login(username)
        login_username = user_obj.username if user_obj else username
        user = authenticate(request, username=login_username, password=password)

        if user is not None:
            _clear_login_failures(throttle_key)
            if _password_expired(user):
                _audit(request, "login_password_expired", user, {"username": username})
                messages.error(
                    request,
                    "La contraseña venció. Solicite a un administrador actualizarla."
                )
                return render(request, "login.html", {
                    "next": request.GET.get("next", ""),
                })

            login(request, user)
            _audit(request, "login_success", user, {"username": username})
            next_url = _safe_next_url(request)
            return redirect(next_url or "dashboard")
        else:
            _audit(request, "login_failed", metadata={"username": username})
            locked = _record_login_failure(throttle_key)
            if locked:
                _audit(request, "login_rate_limited", metadata={"username": username})
                messages.error(
                    request,
                    "Demasiados intentos fallidos. Espere 15 minutos antes de volver a intentar."
                )
                return render(request, "login.html", {
                    "next": request.GET.get("next", ""),
                }, status=429)
            else:
                messages.error(request, "Usuario o contraseña incorrectos")

    return render(request, "login.html", {
        "next": request.GET.get("next", ""),
    })


@login_required(login_url="login")
@require_POST
def logout_view(request):
    if puede_gestionar_horarios(request.user):
        clases = ClaseHorario.objects.filter(archivado=False).exclude(
            profesor=PROFESOR_TALLERES_PRECARGADOS
        )
        if not es_administrador(request.user):
            clases = clases.filter(profesor=request.user.username)
        _cerrar_talleres_terminados(request, clases)

    _audit(request, "logout")
    logout(request)
    return redirect("login")


def _fin_clase_datetime(clase):
    if not clase.fecha or not clase.hora_fin:
        return None

    fecha_fin = clase.fecha
    if clase.hora_inicio and clase.hora_fin < clase.hora_inicio:
        fecha_fin += timedelta(days=1)

    fin = datetime.combine(fecha_fin, clase.hora_fin)
    if timezone.is_naive(fin):
        fin = timezone.make_aware(fin, timezone.get_current_timezone())
    return fin


def _inicio_clase_datetime(clase):
    if not clase.fecha or not clase.hora_inicio:
        return None

    inicio = datetime.combine(clase.fecha, clase.hora_inicio)
    if timezone.is_naive(inicio):
        inicio = timezone.make_aware(inicio, timezone.get_current_timezone())
    return inicio


def _clase_ya_termino(clase, ahora=None):
    fin = _fin_clase_datetime(clase)
    if fin is None:
        return False
    return fin < (ahora or timezone.localtime())


def _ids_clases_pasadas(clases, ahora=None):
    ahora = ahora or timezone.localtime()
    return [clase.pk for clase in clases if _clase_ya_termino(clase, ahora)]


def _actualizar_respaldo_unico(request, motivo, clases=None):
    if is_sqlite_memory_database():
        return None

    destino = create_rolling_backup()
    metadata = {
        "path": str(destino),
        "verified": True,
        "modo": "unico_actualizable",
        "motivo": motivo,
    }
    if clases:
        metadata["talleres"] = [clase.pk for clase in clases]
    _audit(request, "database_rolling_backup_updated", metadata=metadata)
    return destino


def _respaldar_clases_por_finalizar(request, clases=None, ahora=None):
    ahora = ahora or timezone.localtime()
    segundos = max(int(getattr(settings, "CLASS_BACKUP_LEAD_SECONDS", 30)), 1)
    anticipacion = timedelta(seconds=segundos)
    queryset = clases if clases is not None else ClaseHorario.objects.filter(archivado=False)
    queryset = queryset.exclude(profesor=PROFESOR_TALLERES_PRECARGADOS)

    por_respaldar = []
    for clase in list(queryset.filter(respaldo_pre_cierre_at__isnull=True)):
        fin = _fin_clase_datetime(clase)
        if fin is not None and ahora >= fin - anticipacion:
            por_respaldar.append(clase)

    if not por_respaldar:
        return None

    destino = _actualizar_respaldo_unico(
        request,
        "pre_cierre_taller",
        clases=por_respaldar,
    )
    ClaseHorario.objects.filter(
        pk__in=[clase.pk for clase in por_respaldar],
        respaldo_pre_cierre_at__isnull=True,
    ).update(respaldo_pre_cierre_at=ahora)
    return destino


def _cerrar_talleres_terminados(request, clases=None):
    ahora = timezone.localtime()
    queryset = clases if clases is not None else ClaseHorario.objects.all()
    queryset = queryset.exclude(profesor=PROFESOR_TALLERES_PRECARGADOS)
    try:
        _respaldar_clases_por_finalizar(request, queryset, ahora=ahora)
    except (OSError, ValueError) as exc:
        logger.exception("No se pudo actualizar el respaldo previo al cierre: %s", exc)
        _audit(request, "database_rolling_backup_failed", metadata={
            "motivo": "pre_cierre_taller",
            "error": str(exc)[:500],
        })

    cerradas = 0
    for clase in list(queryset):
        if not _clase_ya_termino(clase, ahora):
            continue

        with transaction.atomic():
            clase_actual = ClaseHorario.objects.select_for_update().get(pk=clase.pk)
            if clase_actual.archivado or not _clase_ya_termino(clase_actual, ahora):
                continue
            clase_actual.archivado = True
            clase_actual.save(update_fields=["archivado"])

        cerradas += 1
        _audit(request, "clase_horario_auto_closed", clase_actual)
        try:
            _generar_exportaciones_taller(clase_actual, request=request)
        except Exception as exc:
            _registrar_error_exportacion(clase_actual, request, exc)

    return cerradas


def _asegurar_talleres_precargados(fecha_referencia=None):
    hoy = fecha_referencia or timezone.localdate()
    inicio_semana = hoy - timedelta(days=hoy.weekday())

    for dia_semana in range(5):
        fecha = inicio_semana + timedelta(days=dia_semana)
        for nombre, hora_inicio, hora_fin in TALLERES_PRECARGADOS:
            ClaseHorario.objects.update_or_create(
                fecha=fecha,
                profesor=PROFESOR_TALLERES_PRECARGADOS,
                asignatura=nombre,
                seccion=nombre,
                defaults={
                    "horario": f"{hora_inicio:%H:%M} - {hora_fin:%H:%M}",
                    "dia_semana": dia_semana,
                    "hora_inicio": hora_inicio,
                    "hora_fin": hora_fin,
                    "archivado": False,
                },
            )


def _hora_en_rango(hora_actual, hora_inicio, hora_fin):
    if isinstance(hora_actual, datetime):
        hora_actual = hora_actual.time()
    if getattr(hora_actual, "tzinfo", None) is not None:
        hora_actual = hora_actual.replace(tzinfo=None)

    if hora_inicio <= hora_fin:
        return hora_inicio <= hora_actual <= hora_fin
    return hora_actual >= hora_inicio or hora_actual <= hora_fin


def _talleres_precargados_opciones(ahora=None):
    ahora = ahora or timezone.localtime()
    fecha = ahora.date()
    _asegurar_talleres_precargados(fecha)

    talleres_guardados = ClaseHorario.objects.filter(
        profesor=PROFESOR_TALLERES_PRECARGADOS,
        archivado=False,
    ).order_by("-fecha", "-id")

    # Los talleres precargados son plantillas de horario. Se toma la versión
    # más reciente de cada nombre desde la base de datos para que sigan
    # disponibles también los fines de semana o cuando no exista una fila
    # específica para la fecha actual.
    talleres_por_nombre = {}
    for taller in talleres_guardados:
        talleres_por_nombre.setdefault(taller.asignatura, taller)
    talleres = sorted(
        talleres_por_nombre.values(),
        key=lambda taller: (taller.hora_inicio, taller.asignatura),
    )
    dias_semana = dict(ClaseHorario.DIA_SEMANA_CHOICES)

    return [
        {
            "nombre": taller.asignatura,
            "fecha": fecha,
            "dia_semana": dias_semana[fecha.weekday()],
            "hora_inicio": taller.hora_inicio,
            "hora_fin": taller.hora_fin,
            "seccion": taller.seccion,
            "activo": _hora_en_rango(
                ahora.time(),
                taller.hora_inicio,
                taller.hora_fin,
            ),
        }
        for taller in talleres
    ]


def _crear_taller_precargado_profesor(profesor_nombre, nombre_taller, ahora=None):
    ahora = ahora or timezone.localtime()
    _asegurar_talleres_precargados(ahora.date())

    taller = ClaseHorario.objects.filter(
        profesor=PROFESOR_TALLERES_PRECARGADOS,
        asignatura=nombre_taller,
        archivado=False,
    ).order_by("-fecha", "-id").first()
    if taller is None:
        raise ValueError("Taller precargado no válido.")

    if not _hora_en_rango(ahora.time(), taller.hora_inicio, taller.hora_fin):
        raise ValueError("Este taller solo se puede cargar dentro de su horario.")

    return ClaseHorario.objects.get_or_create(
        profesor=profesor_nombre,
        fecha=ahora.date(),
        asignatura=taller.asignatura,
        seccion=taller.seccion,
        hora_inicio=taller.hora_inicio,
        hora_fin=taller.hora_fin,
        archivado=False,
        defaults={
            "horario": taller.horario,
            "dia_semana": ahora.date().weekday(),
        },
    )


@login_required(login_url="login")
def profesor(request):
    if not puede_gestionar_horarios(request.user):
        messages.error(request, "No tiene permisos para gestionar talleres.")
        return redirect("dashboard")

    profesor_nombre = request.user.username
    _cerrar_talleres_terminados(
        request,
        ClaseHorario.objects.filter(profesor=profesor_nombre, archivado=False),
    )

    total_clases = ClaseHorario.objects.filter(
        profesor=profesor_nombre,
        archivado=False,
    ).count()

    form_abierto = False
    if request.method == 'POST':
        if not puede_crear_talleres(request.user):
            messages.error(request, "Solo profesores y administradores autorizados pueden crear talleres.")
            return redirect('profesor')

        taller_precargado = request.POST.get("taller_precargado")
        if taller_precargado:
            try:
                clase, created = _crear_taller_precargado_profesor(
                    profesor_nombre,
                    taller_precargado,
                )
            except ValueError as error:
                messages.error(request, str(error))
                return redirect("profesor")
            else:
                if created:
                    _audit(
                        request,
                        "clase_horario_created",
                        clase,
                        {"origen": "taller_precargado"},
                    )
                    messages.success(request, f"{clase.asignatura} cargado correctamente.")
                else:
                    messages.info(request, f"{clase.asignatura} ya estaba cargado para hoy.")
                return redirect("registrar_evento")

        form = ClaseHorarioForm(request.POST)
        form_abierto = True

        if form.is_valid():
            nueva_clase = form.save(commit=False)
            nueva_clase.profesor = profesor_nombre
            fecha = form.cleaned_data['fecha']
            nueva_clase.fecha = fecha
            dia_semana = fecha.weekday()   # Lunes=0 ... Domingo=6
            nueva_clase.dia_semana = dia_semana

            # Guardar formato horario "HH:MM - HH:MM"
            nueva_clase.horario = f"{nueva_clase.hora_inicio} - {nueva_clase.hora_fin}"

            nueva_clase.save()
            _audit(
                request,
                "clase_horario_created",
                nueva_clase,
                {"tipo_taller": nueva_clase.tipo_taller},
            )
            messages.success(
                request,
                f"Taller creado como {nueva_clase.get_tipo_taller_display()}.",
            )
            return redirect('profesor')

    else:
        form = ClaseHorarioForm()

    clases = list(ClaseHorario.objects.filter(
        profesor=profesor_nombre,
        archivado=False,
    ).order_by('fecha', 'dia_semana', 'hora_inicio'))

    actividades_pasadas_count = len(_ids_clases_pasadas(clases))

    context = {
        'form': form,
        'total_clases': total_clases,
        'clases': clases,
        'actividades_pasadas_count': actividades_pasadas_count,
        'tipo_taller_opciones': ClaseHorario.TIPO_TALLER_CHOICES,
        'talleres_opciones': _talleres_precargados_opciones(),
        'form_abierto': form_abierto,
        'puede_crear_taller': puede_crear_talleres(request.user),
    }
    return render(request, 'profesor.html', context)


@login_required(login_url="login")
@require_POST
def archivar_actividades_pasadas(request):
    if not puede_gestionar_horarios(request.user):
        messages.error(request, "No tiene permisos para archivar talleres.")
        return redirect("dashboard")

    clases = ClaseHorario.objects.filter(
        profesor=request.user.username,
        archivado=False,
    )
    actualizadas = _cerrar_talleres_terminados(request, clases)

    _audit(request, "clases_horario_archived", metadata={"total": actualizadas})
    if actualizadas:
        messages.success(request, f"Talleres finalizados archivados: {actualizadas}.")
    else:
        messages.info(request, "No hay talleres finalizados para archivar.")
    return redirect("profesor")


@login_required(login_url="login")
@require_POST
def eliminar_actividad(request, id):
    if not puede_gestionar_horarios(request.user):
        messages.error(request, "No tiene permisos para eliminar talleres.")
        return redirect("dashboard")

    queryset = ClaseHorario.objects.all()
    if not es_administrador(request.user):
        queryset = queryset.filter(profesor=request.user.username)

    clase = get_object_or_404(queryset, pk=id)
    nombre = clase.asignatura
    _audit(request, "clase_horario_deleted", clase)
    clase.delete()

    messages.success(request, f"Taller eliminado: {nombre}.")
    return redirect("profesor")

@login_required(login_url="login")
@require_POST
def actualizar_residuo(request):
    if not es_operador(request.user):
        return JsonResponse({"success": False, "error": "No tiene permisos para actualizar residuos."}, status=403)

    residuo = get_object_or_404(Residuo, id=request.POST.get("residuo_id"))
    if not _puede_gestionar_pendiente(request.user, residuo):
        return JsonResponse({"success": False, "error": "No puede modificar un residuo creado por otro usuario."}, status=403)
    if residuo.estado != ESTADO_RESIDUO_PENDIENTE:
        return JsonResponse({
            "success": False,
            "error": "Este registro ya no está pendiente. Actualice el panel antes de continuar.",
        }, status=409)

    try:
        categoria_id = _parse_int(
            request.POST.get("tipo"),
            "tipo de residuo",
            min_value=1,
        )
        categoria = _categorias_residuo_disponibles().filter(
            id_categoria=categoria_id,
        ).first()
        if categoria is None:
            raise ValueError(
                "El tipo de residuo debe ser Orgánico o Inorgánico."
            )
        subtipo = _subtipo_interno_categoria(categoria)
        tipo_operacional = tipo_operacional_categoria(categoria)

        peso = None
        unidades = None
        asignaciones_organicas = []

        residuo.tipo = categoria
        residuo.subtipo = subtipo
        residuo.peso = None
        residuo.unidad = None
        residuo.contenedor_id = "Sin Asignar"

        if tipo_operacional == TIPO_ORGANICO:
            peso = _parse_decimal(
                request.POST.get("peso"),
                "peso",
                min_value=0,
                max_value=_weight_max(),
            )
            if peso <= 0:
                raise ValueError(
                    "Debe ingresar un peso mayor a 0 para residuos orgánicos."
                )
            asignaciones_organicas = distribuir_peso_organico(
                peso,
                residuo_excluido_id=residuo.pk,
            )
            if not asignaciones_organicas:
                raise ValueError(
                    "El peso ingresado supera el espacio disponible conjunto de "
                    "Compostera 1 y Compostera 2. Compostera 3 está deshabilitada "
                    "para nuevos ingresos."
                )
            residuo.contenedor_id = asignaciones_organicas[0]["contenedor"]
            residuo.peso = asignaciones_organicas[0]["peso_kg"]
        elif tipo_operacional == TIPO_INORGANICO:
            unidades = _parse_int(
                request.POST.get("unidad"),
                "cantidad de unidades",
                min_value=1,
                max_value=LIMITE_UNIDADES_GLOBAL,
            )
            residuo.contenedor_id = "Unidades"
            residuo.unidad = unidades
        else:
            raise ValueError(
                "El tipo de residuo debe ser Orgánico o Inorgánico."
            )

        ahora = timezone.now()
        residuo.hora_escaneo = ahora
        residuo.estado = ESTADO_RESIDUO_CONFIRMADO
        residuo.confirmado_at = ahora
        residuo.updated_by = request.user
        residuo.source_ip = _client_ip(request)
        residuos_adicionales = []
        with transaction.atomic():
            if tipo_operacional == TIPO_INORGANICO:
                _bloquear_capacidad("unidades")
                _validar_capacidad_unidades(
                    unidades,
                    residuo_excluido_id=residuo.pk,
                )
            residuo.save()
            if tipo_operacional == TIPO_ORGANICO and len(asignaciones_organicas) > 1:
                ultimo_numero = (
                    Residuo.objects
                    .filter(asignatura=residuo.asignatura, seccion=residuo.seccion)
                    .aggregate(maximo=Max("numero_clase"))["maximo"] or 0
                )
                for indice, asignacion in enumerate(asignaciones_organicas[1:], start=1):
                    adicional = Residuo.objects.create(
                        taller_id=residuo.taller_id,
                        seccion=residuo.seccion,
                        profesor=residuo.profesor,
                        asignatura=residuo.asignatura,
                        horario=residuo.horario,
                        numero_clase=int(ultimo_numero) + indice,
                        hora_escaneo=ahora,
                        tipo=categoria,
                        subtipo=subtipo,
                        peso=asignacion["peso_kg"],
                        unidad=None,
                        contenedor_id=asignacion["contenedor"],
                        estado=ESTADO_RESIDUO_CONFIRMADO,
                        retirado=False,
                        confirmado_at=ahora,
                        created_by=residuo.created_by or request.user,
                        updated_by=request.user,
                        source_ip=_client_ip(request),
                        manual_entry=residuo.manual_entry,
                        manual_reason=residuo.manual_reason,
                    )
                    residuos_adicionales.append(adicional)

        resumen_contenedores = residuo.contenedor_id
        if tipo_operacional == TIPO_ORGANICO and asignaciones_organicas:
            resumen_contenedores = " y ".join(
                f'{item["contenedor"]} ({item["peso_kg"]} kg)'
                for item in asignaciones_organicas
            )
        elif tipo_operacional == TIPO_INORGANICO:
            resumen_contenedores = f"Unidades ({unidades} unidades)"
        _audit(request, "residuo_updated", residuo, {
            "categoria": categoria.nombre,
            "contenedor": resumen_contenedores,
            "registros_generados": 1 + len(residuos_adicionales),
        })
        for adicional in residuos_adicionales:
            _audit(request, "residuo_split_created", adicional, {
                "residuo_origen": residuo.pk,
                "contenedor": adicional.contenedor_id,
            })

        return JsonResponse({
            "success": True,
            "contenedor_asignado": resumen_contenedores,
            "registro_contexto": {
                "taller_id": residuo.taller_id,
                "seccion": residuo.seccion,
                "profesor": residuo.profesor,
                "asignatura": residuo.asignatura,
                "horario": residuo.horario,
            },
        })
    except ValueError as error:
        return JsonResponse({"success": False, "error": str(error)}, status=400)
    except OperationalError:
        return JsonResponse({
            "success": False,
            "error": "Otro usuario está actualizando la capacidad. Intente nuevamente.",
        }, status=409)




@login_required(login_url="login")
def confirmar_residuo(request):
    """Vista para mostrar el formulario de registro orgánico por peso."""
    if not es_operador(request.user):
        messages.error(request, "No tiene permisos para confirmar residuos.")
        return redirect("dashboard")

    return render(request, "residuo.html")


def _datetime_manual_taller(fecha, taller):
    hora = taller.hora_inicio or time(0, 0)
    fecha_hora = datetime.combine(fecha, hora)
    if timezone.is_naive(fecha_hora):
        fecha_hora = timezone.make_aware(fecha_hora, timezone.get_current_timezone())
    return fecha_hora


def _crear_residuo_manual(request, cleaned_data):
    taller = cleaned_data["taller"]
    categoria = cleaned_data["categoria"]
    tipo_residuo = _subtipo_interno_categoria(categoria)
    tipo_medicion = cleaned_data["tipo_medicion"]
    cantidad = cleaned_data["cantidad"]
    motivo = cleaned_data["motivo"]
    tipo_normalizado = tipo_operacional_categoria(categoria)

    contenedor_id = "Sin Asignar"
    peso = None
    unidad = None
    asignaciones_registro = []

    if tipo_normalizado == TIPO_ORGANICO:
        asignaciones = distribuir_peso_organico(cantidad)
        if not asignaciones:
            raise ValueError(
                "El peso ingresado supera el espacio disponible conjunto de "
                "Compostera 1 y Compostera 2. Compostera 3 está deshabilitada "
                "para nuevos ingresos."
            )
        asignaciones_registro = [
            {
                "contenedor_id": item["contenedor"],
                "peso": item["peso_kg"],
                "unidad": None,
            }
            for item in asignaciones
        ]
    elif tipo_normalizado == TIPO_INORGANICO:
        contenedor_id = "Unidades"
        if tipo_medicion == MEDICION_UNIDADES:
            unidad = int(cantidad)
    else:
        raise ValueError("La categoría debe ser orgánica o inorgánica.")

    if not asignaciones_registro:
        asignaciones_registro = [{
            "contenedor_id": contenedor_id,
            "peso": peso,
            "unidad": unidad,
        }]

    with transaction.atomic():
        if tipo_normalizado == TIPO_INORGANICO:
            _bloquear_capacidad("unidades")
            _validar_capacidad_unidades(unidad)

        ultimo_numero = (
            Residuo.objects
            .filter(asignatura=taller.asignatura, seccion=taller.seccion)
            .aggregate(maximo=Max("numero_clase"))["maximo"] or 0
        )
        numero_clase = int(ultimo_numero) + 1

        residuos = []
        confirmado_at = timezone.now()
        hora_escaneo = _datetime_manual_taller(cleaned_data["fecha"], taller)
        for indice, asignacion in enumerate(asignaciones_registro):
            residuo = Residuo.objects.create(
                taller=taller,
                seccion=taller.seccion,
                profesor=taller.profesor,
                asignatura=taller.asignatura,
                horario=taller.horario,
                numero_clase=numero_clase + indice,
                hora_escaneo=hora_escaneo,
                tipo=categoria,
                subtipo=tipo_residuo,
                peso=asignacion["peso"],
                unidad=asignacion["unidad"],
                contenedor_id=asignacion["contenedor_id"],
                estado=ESTADO_RESIDUO_CONFIRMADO,
                retirado=False,
                confirmado_at=confirmado_at,
                created_by=request.user,
                updated_by=request.user,
                source_ip=_client_ip(request),
                manual_entry=True,
                manual_reason=motivo,
            )
            residuos.append(residuo)

    resumen_contenedores = " y ".join(
        f'{item["contenedor_id"]} ({item["peso"]} kg)'
        if item["peso"] is not None else item["contenedor_id"]
        for item in asignaciones_registro
    )
    for residuo in residuos:
        _audit(request, "residuo_manual_created", residuo, {
            "taller_id": taller.pk,
            "motivo": motivo,
            "contenedor": residuo.contenedor_id,
            "tipo_medicion": tipo_medicion,
            "registros_generados": len(residuos),
        })

    if taller.archivado:
        try:
            _generar_exportaciones_taller(taller, request=request, force=True)
        except Exception as exc:
            _registrar_error_exportacion(taller, request, exc)

    return residuos[0], resumen_contenedores


@manual_entry_required
def ingreso_manual_olvido(request):
    asegurar_tipos_residuos_predeterminados()
    if request.method == "POST":
        form = ManualResiduoOlvidoForm(request.POST, user=request.user)
        if form.is_valid():
            try:
                residuo, resumen_contenedores = _crear_residuo_manual(
                    request,
                    form.cleaned_data,
                )
            except ValueError as error:
                form.add_error(None, str(error))
            except OperationalError:
                form.add_error(
                    None,
                    "Otro usuario está actualizando la capacidad. Intente nuevamente.",
                )
            else:
                try:
                    _actualizar_respaldo_unico(
                        request,
                        "ingreso_manual",
                        clases=[residuo.taller] if residuo.taller_id else None,
                    )
                except (OSError, ValueError) as exc:
                    logger.exception(
                        "El ingreso manual se guardó, pero falló el respaldo inmediato: %s",
                        exc,
                    )
                    _audit(request, "database_rolling_backup_failed", metadata={
                        "motivo": "ingreso_manual",
                        "residuo_id": residuo.pk,
                        "error": str(exc)[:500],
                    })
                    messages.warning(
                        request,
                        "El ingreso quedó registrado, pero no se pudo actualizar el respaldo.",
                    )
                messages.success(
                    request,
                    f"Ingreso manual registrado en {resumen_contenedores}.",
                )
                return redirect("dashboard")
    else:
        form = ManualResiduoOlvidoForm(user=request.user)

    return render(request, "ingreso_manual.html", {"form": form})

@login_required(login_url="login")
@require_POST
def receive_weight(request):
    try:
        data = json.loads(request.body.decode("utf-8"))
        reading = _crear_lectura_peso(request, data, user=request.user)
        return JsonResponse({"ok": True, "id": reading.id, "record_code": str(reading.record_code)})
    except (json.JSONDecodeError, UnicodeDecodeError):
        return JsonResponse({"ok": False, "error": "JSON inválido."}, status=400)
    except ValueError as error:
        return JsonResponse({"ok": False, "error": str(error)}, status=400)


@csrf_exempt
@require_POST
def receive_weight_api(request):
    expected_token = settings.WEIGHT_API_TOKEN
    received_token = (
        request.headers.get("X-Weight-Token")
        or request.headers.get("Authorization", "").removeprefix("Bearer ").strip()
    )

    if not expected_token:
        return JsonResponse({
            "ok": False,
            "error": "WEIGHT_API_TOKEN no está configurado en el entorno."
        }, status=503)

    if not received_token or not secrets.compare_digest(received_token, expected_token):
        _audit(request, "weight_api_token_failed")
        return JsonResponse({"ok": False, "error": "Token inválido."}, status=403)

    try:
        data = json.loads(request.body.decode("utf-8"))
        reading = _crear_lectura_peso(request, data)
        return JsonResponse({"ok": True, "id": reading.id, "record_code": str(reading.record_code)})
    except (json.JSONDecodeError, UnicodeDecodeError):
        return JsonResponse({"ok": False, "error": "JSON inválido."}, status=400)
    except ValueError as error:
        return JsonResponse({"ok": False, "error": str(error)}, status=400)


@login_required(login_url="login")
@require_GET
def balanza_leer(request):
    if not es_operador(request.user):
        return JsonResponse({
            "ok": False,
            "peso": None,
            "mensaje": "No tiene permisos para leer la balanza.",
        }, status=403)

    if getattr(settings, "WEIGHT_BRIDGE_FIRST", True):
        lectura_reciente = _ultima_lectura_peso_reciente()
        if lectura_reciente:
            return JsonResponse(_lectura_peso_json(lectura_reciente))

    if not getattr(settings, "WEIGHT_DIRECT_READ_ENABLED", True):
        return JsonResponse(_error_balanza_json({
            "codigo": "puente_sin_lectura",
            "mensaje": "No hay una lectura reciente del puente de balanza.",
            "accion": "Inicie scripts\\weight_bridge.ps1 y vuelva a intentar.",
        }))

    resultado = balanza_service.obtener_peso_estable()
    if not resultado.get("ok"):
        logger.warning("Búsqueda automática de balanza sin lectura: %s", resultado)
        lectura_reciente = _ultima_lectura_peso_reciente()
        if lectura_reciente:
            return JsonResponse(_lectura_peso_json(lectura_reciente))
        return JsonResponse(_error_balanza_json(resultado))

    reading = _crear_lectura_peso(request, {
        "weight_kg": str(resultado["peso"]),
        "device_name": resultado.get("dispositivo") or resultado.get("puerto") or "balanza-automatica",
        "raw_data": resultado.get("raw") or "",
        "is_stable": True,
    }, user=request.user)

    return JsonResponse({
        "ok": True,
        "peso": float(reading.weight_kg),
        "unidad": "kg",
        "puerto": resultado.get("puerto"),
        "baudrate": resultado.get("baudrate"),
        "serial_mode": resultado.get("serial_mode"),
        "line_control": resultado.get("line_control"),
        "mensaje": "Balanza detectada y lectura estable.",
        "fuente": resultado.get("fuente") or "serial",
        "dispositivo": reading.device_name,
        "puertos_probados": resultado.get("puertos_probados", []),
    })


@login_required(login_url="login")
@require_GET
def balanza_diagnostico(request):
    if not es_operador(request.user):
        return JsonResponse(
            {"ok": False, "mensaje": "No tiene permisos para diagnosticar la balanza."},
            status=403,
        )
    payload = balanza_service.diagnostico_balanza()
    lectura_reciente = _ultima_lectura_peso_reciente()
    payload["lectura_reciente"] = (
        _lectura_peso_json(lectura_reciente) if lectura_reciente else None
    )
    return JsonResponse(payload)


@login_required(login_url="login")
@require_GET
def last_weight(request):
    reading = _ultima_lectura_peso_reciente()

    if reading:
        return JsonResponse({
            "weight_kg": float(reading.weight_kg),
            "device_name": reading.device_name,
            "is_stable": reading.is_stable,
            "created_at": timezone.localtime(reading.created_at).isoformat(),
        })

    return JsonResponse({"weight_kg": None})


@admin_required
def usuarios_view(request):
    if request.method == "POST":
        form = UsuariosForm(request.POST, require_password=True)
        if form.is_valid():
            try:
                with transaction.atomic():
                    usuario = form.save(commit=False)
                    usuario.email = form.cleaned_data["email"]
                    usuario.save()
                    _sincronizar_auth_user(
                        usuario,
                        password=form.cleaned_data["password1"],
                        grupo=form.cleaned_data["grupo"]
                    )
                _audit(request, "usuario_created", usuario, {"email": usuario.email})
                try:
                    enviar_confirmacion_registro_usuario(
                        usuario,
                        request.build_absolute_uri(reverse("login")),
                    )
                except Exception as exc:
                    logger.exception(
                        "No se pudo enviar el correo de confirmación a %s.",
                        usuario.email,
                    )
                    _audit(
                        request,
                        "usuario_confirmation_email_failed",
                        usuario,
                        {"error": str(exc)[:500]},
                    )
                    messages.warning(
                        request,
                        "Usuario creado correctamente, pero no se pudo enviar el correo de confirmación.",
                    )
                else:
                    _audit(
                        request,
                        "usuario_confirmation_email_sent",
                        usuario,
                        {"email": usuario.email},
                    )
                    messages.success(
                        request,
                        "Usuario creado correctamente. Se envió un correo de confirmación.",
                    )
                return redirect("usuarios")
            except IntegrityError:
                messages.error(request, "No se pudo crear el usuario porque el correo ya existe.")
    else:
        form = UsuariosForm(require_password=True)

    listadoUsuarios = Usuarios.objects.select_related("user").all()
    return render(request, "usuarios.html", {
        "form": form,
        "Usuarios": listadoUsuarios
    })


@login_required(login_url="login")
def base_view(request):
    return redirect("dashboard")


@admin_required
@require_POST
def eliminarUsuario(request, id_usuario):
    usuario = get_object_or_404(Usuarios, id_usuario=id_usuario)
    if usuario.user_id == request.user.id:
        messages.error(request, "No puede eliminar su propio usuario mientras está conectado.")
        return redirect("usuarios")

    email = usuario.email
    user = usuario.user
    if user:
        user.delete()
    else:
        usuario.delete()
    _audit(request, "usuario_deleted", metadata={"email": email})
    messages.success(request, "Usuario eliminado correctamente.")
    return redirect("usuarios")


@admin_required
def editarUsuario(request, id_usuario):
    usuario = get_object_or_404(Usuarios, id_usuario=id_usuario)
    user = usuario.user

    if request.method == "POST":
        form = UsuariosForm(
            request.POST,
            instance=usuario,
            require_password=False,
            current_user=user
        )
        if form.is_valid():
            try:
                with transaction.atomic():
                    usuario = form.save(commit=False)
                    usuario.email = form.cleaned_data["email"]
                    usuario.save()
                    _sincronizar_auth_user(
                        usuario,
                        password=form.cleaned_data.get("password1"),
                        grupo=form.cleaned_data["grupo"]
                    )
                _audit(request, "usuario_updated", usuario, {"email": usuario.email})
                messages.success(request, "Usuario actualizado correctamente.")
                return redirect("usuarios")
            except IntegrityError:
                messages.error(request, "No se pudo actualizar el usuario porque el correo ya existe.")
    else:
        form = UsuariosForm(instance=usuario, require_password=False, current_user=user)

    return render(request, "editar_usuario.html", {
        "form": form,
        "id_usuario": id_usuario,
        "usuario_editado": usuario,
    })


@admin_required
def gestionCategorias(request):
    asegurar_tipos_residuos_predeterminados()
    form = CategoriaResiduosForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        categoria = form.save()
        _audit(request, "categoria_created", categoria)
        messages.success(request, f"Categoría agregada: {categoria.nombre}.")
        return redirect("gestionCategorias")

    categorias = _categorias_residuo_disponibles()
    return render(request, 'gestionCategorias.html', {
        'form': form,
        'categoriaResiduos': categorias,
    })
@admin_required
@require_POST
def eliminarCategoria(request, id_categoria):
    categoria = get_object_or_404(CategoriaResiduos, id_categoria=id_categoria)
    nombre = categoria.nombre
    categoria.delete()
    _audit(request, "categoria_deleted", metadata={"nombre": nombre})
    return redirect("gestionCategorias")

@admin_required
def editarCategoria(request, id_categoria):
    categoria = get_object_or_404(CategoriaResiduos, id_categoria=id_categoria)

    if request.method == "POST":
        form = CategoriaResiduosForm(request.POST, instance=categoria)
        if form.is_valid():
            categoria = form.save()
            _audit(request, "categoria_updated", categoria)
            return redirect("gestionCategorias")
    else:
        form = CategoriaResiduosForm(instance=categoria)

    return render(request, "editarCategoria.html", {
        "form": form,
        "id_categoria": id_categoria
    })
@login_required(login_url="login")
def dashboard(request):
    if puede_gestionar_horarios(request.user):
        clases = ClaseHorario.objects.filter(archivado=False)
        if not es_administrador(request.user):
            clases = clases.filter(profesor=request.user.username)
        _cerrar_talleres_terminados(request, clases)

    context = crear_dashboard_context(request.user)
    return render(request, 'dashboard.html', context)


@login_required(login_url="login")
@require_GET
def dashboard_unidades(request):
    return JsonResponse(obtener_datos_unidades())


@login_required(login_url="login")
@require_POST
def cerrar_residuo_pendiente(request, residuo_id):
    if not es_operador(request.user):
        messages.error(request, "No tiene permisos para cerrar registros pendientes.")
        return redirect("dashboard")

    with transaction.atomic():
        residuo = get_object_or_404(
            Residuo.objects.select_for_update(),
            pk=residuo_id,
        )
        if not _puede_gestionar_pendiente(request.user, residuo):
            messages.error(request, "No puede cerrar un registro pendiente creado por otro usuario.")
            return redirect("dashboard")
        if residuo.estado != ESTADO_RESIDUO_PENDIENTE:
            messages.info(request, "El registro seleccionado ya no estaba pendiente.")
            return redirect("dashboard")

        residuo.estado = ESTADO_RESIDUO_ANULADO
        residuo.retirado = True
        residuo.updated_by = request.user
        residuo.save(update_fields=["estado", "retirado", "updated_by"])
        _audit(request, "residuo_pending_closed", residuo, {
            "taller_id": residuo.taller_id,
        })

    messages.success(request, f"Pendiente #{residuo.id} cerrado correctamente.")
    return redirect("dashboard")


@admin_required
@require_POST
def limpiar_panel(request):
    with transaction.atomic():
        _bloquear_capacidad("unidades")
        residuos_activos = Residuo.objects.filter(
            estado=ESTADO_RESIDUO_CONFIRMADO,
            retirado=False,
        )
        activos_actualizados = residuos_activos.update(
            retirado=True,
            updated_by=request.user,
        )
        pendientes_actualizados = Residuo.objects.filter(
            estado=ESTADO_RESIDUO_PENDIENTE,
        ).update(
            estado=ESTADO_RESIDUO_ANULADO,
            retirado=True,
            updated_by=request.user,
        )

        _audit(
            request,
            "dashboard_cleared",
            metadata={
                "residuos_activos_limpiados": activos_actualizados,
                "residuos_pendientes_anulados": pendientes_actualizados,
            },
        )

    messages.success(
        request,
        f"Panel limpiado: {activos_actualizados} registro(s) activo(s) y "
        f"{pendientes_actualizados} pendiente(s) fuera del panel.",
    )
    return redirect("dashboard")


@login_required(login_url="login")
@require_POST
def registrar_retiro(request):
    if not es_operador(request.user):
        return JsonResponse({"success": False, "msg": "No tiene permisos para registrar retiros."}, status=403)

    try:
        origen = _clean_text(request.POST.get("origen"), 100)
        cantidad = _parse_decimal(request.POST.get("cantidad"), "cantidad", min_value=0, max_value=_weight_max(), required=False) or Decimal("0")
        destino_obj = Destino.objects.get(id=request.POST.get("destino"), activo=True)
        cat_obj = CategoriaResiduos.objects.get(id_categoria=request.POST.get("categoria"))
        tipo_obj = TipoResiduos.objects.get(id_tipo=request.POST.get("subtipo"), categoria=cat_obj)
        subtipo_nombre = tipo_obj.nombre_residuo or "general"
        tipo_normalizado = tipo_operacional_categoria(cat_obj)
    except (Destino.DoesNotExist, CategoriaResiduos.DoesNotExist, TipoResiduos.DoesNotExist):
        return JsonResponse({"success": False, "msg": "Datos de retiro inválidos o destino inactivo."}, status=400)
    except ValueError as error:
        return JsonResponse({"success": False, "msg": str(error)}, status=400)
    except OperationalError:
        return JsonResponse({
            "success": False,
            "msg": "Otro usuario está actualizando la capacidad. Intente nuevamente.",
        }, status=409)

    try:
        with transaction.atomic():
            if tipo_normalizado == TIPO_INORGANICO:
                _bloquear_capacidad("unidades")
                residuos_a_retirar = Residuo.objects.filter(
                    contenedor_id="Unidades",
                    tipo=cat_obj,
                    retirado=False,
                    estado=ESTADO_RESIDUO_CONFIRMADO
                )
                total_unidades = int(residuos_a_retirar.aggregate(sum=Sum("unidad"))["sum"] or 0)

                if total_unidades <= 0:
                    return JsonResponse({"success": False, "msg": "No hay unidades para retirar."})

                cantidad_retirar = total_unidades if cantidad <= 0 else int(min(Decimal(total_unidades), cantidad))
                sobrante = total_unidades - cantidad_retirar

                historial = HistorialRetiro.objects.create(
                    usuario=request.user,
                    contenedor_origen="Unidades",
                    tipo_residuo="inorganico",
                    destino=destino_obj,
                    cantidad_peso=0,
                    cantidad_unidades=cantidad_retirar,
                    detalle=f"Reciclaje de {cantidad_retirar} unidades ({subtipo_nombre}). Destino: {destino_obj.nombre}"
                )

                residuos_a_retirar.update(retirado=True, updated_by=request.user)

                if sobrante > 0:
                    Residuo.objects.create(
                        contenedor_id="Unidades",
                        tipo=cat_obj,
                        subtipo=tipo_obj,
                        unidad=sobrante,
                        peso=None,
                        retirado=False,
                        estado=ESTADO_RESIDUO_CONFIRMADO,
                        confirmado_at=timezone.now(),
                        seccion="N/A",
                        profesor="N/A",
                        asignatura="N/A",
                        horario="N/A",
                        numero_clase=None,
                        created_by=request.user,
                        updated_by=request.user,
                        source_ip=_client_ip(request),
                    )

                _audit(request, "retiro_created", historial, {"tipo": "inorganico", "cantidad": cantidad_retirar})
                return JsonResponse({"success": True, "msg": f"Se retiraron {cantidad_retirar} unidades. Quedan {sobrante}."})

            if tipo_normalizado == TIPO_ORGANICO:
                residuos = Residuo.objects.filter(
                    contenedor_id=origen,
                    tipo=cat_obj,
                    retirado=False,
                    estado=ESTADO_RESIDUO_CONFIRMADO
                )
                total_peso = residuos.aggregate(sum=Sum("peso"))["sum"] or Decimal("0")

                if total_peso <= 0:
                    return JsonResponse({"success": False, "msg": "No hay peso orgánico para retirar."})

                cantidad_retirar = total_peso if cantidad <= 0 else min(total_peso, cantidad)
                sobrante = total_peso - cantidad_retirar
                if origen == "Compostera 3" and sobrante > 0:
                    return JsonResponse({
                        "success": False,
                        "msg": "Compostera 3 está deshabilitada; solo se permite retirar su contenido completo sin generar sobrantes.",
                    }, status=400)

                historial = HistorialRetiro.objects.create(
                    usuario=request.user,
                    contenedor_origen=origen,
                    tipo_residuo="organico",
                    destino=destino_obj,
                    cantidad_peso=cantidad_retirar,
                    cantidad_unidades=0,
                    detalle=f"Cosecha de compost ({subtipo_nombre}). Destino: {destino_obj.nombre}"
                )

                residuos.update(retirado=True, updated_by=request.user)

                if sobrante > 0:
                    Residuo.objects.create(
                        contenedor_id=origen,
                        tipo=cat_obj,
                        subtipo=tipo_obj,
                        peso=sobrante,
                        unidad=0,
                        retirado=False,
                        estado=ESTADO_RESIDUO_CONFIRMADO,
                        confirmado_at=timezone.now(),
                        seccion="N/A",
                        profesor="N/A",
                        asignatura="N/A",
                        horario="N/A",
                        numero_clase=None,
                        created_by=request.user,
                        updated_by=request.user,
                        source_ip=_client_ip(request),
                    )

                _audit(request, "retiro_created", historial, {"tipo": "organico", "cantidad": str(cantidad_retirar)})
                return JsonResponse({"success": True, "msg": f"Se retiraron {cantidad_retirar:.3f} kg. Quedan {sobrante:.3f} kg."})

        return JsonResponse({"success": False, "msg": "Categoría de retiro no reconocida."}, status=400)
    except ValueError as error:
        return JsonResponse({"success": False, "msg": str(error)}, status=400)
    except OperationalError:
        return JsonResponse({
            "success": False,
            "msg": "Otro usuario está actualizando la capacidad. Intente nuevamente.",
        }, status=409)


@admin_required
def destinos_list_create(request, destino_id=None):
    destinos = Destino.objects.filter(
        categoria__tipo_operacional__in=(TIPO_ORGANICO, TIPO_INORGANICO),
    ).order_by("id", "nombre")

    destino_edit = None
    if destino_id:
        destino_edit = get_object_or_404(Destino, id=destino_id)

    if request.method == "POST":
        if destino_edit:
            form = DestinoForm(request.POST, instance=destino_edit)
        else:
            form = DestinoForm(request.POST)

        if form.is_valid():
            destino = form.save()
            _audit(request, "destino_updated" if destino_edit else "destino_created", destino)
            messages.success(request, "Destino guardado correctamente.")
            return redirect("destinos")
    else:
        form = DestinoForm(instance=destino_edit)

    return render(request, "destinos.html", {
        "form": form,
        "destinos": destinos,
        "destino_edit": destino_edit,
    })

def normalizar_tipo(tipo: str) -> str:
    return normalizar_tipo_operacional(tipo)

@login_required(login_url="login")
@require_GET
def destinos_por_tipo(request):
    tipo_in = request.GET.get("tipo", "")
    tipo = normalizar_tipo(tipo_in)

    if not tipo:
        return JsonResponse({"ok": False, "error": "Tipo inválido", "tipo_recibido": tipo_in}, status=400)

    cats = CategoriaResiduos.objects.filter(tipo_operacional=tipo)

    cat_ids = list(cats.values_list("id_categoria", flat=True))

    destinos = (
        Destino.objects
        .filter(activo=True, categoria_id__in=cat_ids)
        .values("id", "nombre")
        .order_by("nombre")
    )

    return JsonResponse({
        "ok": True,
        "tipo": tipo,
        "categorias_ids": cat_ids,
        "destinos": list(destinos),
    })



@admin_required
@require_POST
def destino_eliminar(request, id):
    destino = get_object_or_404(Destino, id=id)
    nombre = destino.nombre
    destino.delete()
    _audit(request, "destino_deleted", metadata={"nombre": nombre})
    messages.success(request, "Destino eliminado correctamente.")
    return redirect("destinos")


def _xlsx_col_name(index):
    name = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name


def _xlsx_cell(col_index, row_index, value, style_id=0):
    cell_ref = f"{_xlsx_col_name(col_index)}{row_index}"

    if isinstance(value, Decimal):
        return f'<c r="{cell_ref}" s="{style_id}"><v>{format(value, "f")}</v></c>'
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return f'<c r="{cell_ref}" s="{style_id}"><v>{value}</v></c>'

    text = "" if value is None else str(value)
    preserve = ' xml:space="preserve"' if text != text.strip() or "\n" in text else ""
    return (
        f'<c r="{cell_ref}" s="{style_id}" t="inlineStr">'
        f"<is><t{preserve}>{xml_escape(text)}</t></is>"
        f"</c>"
    )


def _xlsx_row(row_index, values, default_style=0):
    cells = []
    for col_index, item in enumerate(values, start=1):
        value = item
        style_id = default_style
        if isinstance(item, tuple) and len(item) == 2:
            value, style_id = item
        cells.append(_xlsx_cell(col_index, row_index, value, style_id))
    return f'<row r="{row_index}">{"".join(cells)}</row>'


def _format_report_datetime(value):
    if not value:
        return ""
    return timezone.localtime(value).strftime("%d-%m-%Y %H:%M:%S")


def _html_excel_report_response(filename, title, columns, rows):
    data_rows = list(rows)
    generated_at = timezone.localtime().strftime("%d-%m-%Y %H:%M:%S")

    header_cells = "".join(f"<th>{xml_escape(str(column))}</th>" for column in columns)
    body_rows = []
    for row in data_rows:
        padded_row = list(row) + [""] * (len(columns) - len(row))
        cells = "".join(
            f"<td>{xml_escape('' if value is None else str(value))}</td>"
            for value in padded_row[:len(columns)]
        )
        body_rows.append(f"<tr>{cells}</tr>")

    if not body_rows:
        body_rows.append(
            f'<tr><td colspan="{len(columns)}" class="empty">Sin registros para mostrar.</td></tr>'
        )

    html = f"""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <style>
    body {{ font-family: Calibri, Arial, sans-serif; color: #111827; }}
    .report-title {{ background: #0891b2; color: #ffffff; font-size: 20px; font-weight: 700; padding: 12px; }}
    .meta {{ margin: 10px 0 16px; font-size: 12px; color: #475569; }}
    table {{ border-collapse: collapse; width: 100%; }}
    th {{ background: #0f172a; color: #ffffff; border: 1px solid #cbd5e1; padding: 8px; text-align: left; font-weight: 700; }}
    td {{ border: 1px solid #cbd5e1; padding: 7px; vertical-align: top; mso-number-format: "\\@"; }}
    tr:nth-child(even) td {{ background: #f8fafc; }}
    .empty {{ text-align: center; color: #64748b; font-style: italic; }}
  </style>
</head>
<body>
  <div class="report-title">{xml_escape(title)}</div>
  <div class="meta">
    Generado: {xml_escape(generated_at)}<br>
    Total registros: {len(data_rows)}
  </div>
  <table>
    <thead><tr>{header_cells}</tr></thead>
    <tbody>{"".join(body_rows)}</tbody>
  </table>
</body>
</html>"""

    response = HttpResponse(html, content_type="application/vnd.ms-excel; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _xlsx_styles_xml():
    return """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <fonts count="4">
    <font><sz val="11"/><color theme="1"/><name val="Calibri"/></font>
    <font><b/><sz val="16"/><color rgb="FFFFFFFF"/><name val="Calibri"/></font>
    <font><b/><sz val="11"/><color rgb="FFFFFFFF"/><name val="Calibri"/></font>
    <font><b/><sz val="11"/><color rgb="FF334155"/><name val="Calibri"/></font>
  </fonts>
  <fills count="5">
    <fill><patternFill patternType="none"/></fill>
    <fill><patternFill patternType="gray125"/></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FF0891B2"/><bgColor indexed="64"/></patternFill></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FF0F172A"/><bgColor indexed="64"/></patternFill></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FFEFF6FF"/><bgColor indexed="64"/></patternFill></fill>
  </fills>
  <borders count="2">
    <border><left/><right/><top/><bottom/><diagonal/></border>
    <border>
      <left style="thin"><color rgb="FFD1D5DB"/></left>
      <right style="thin"><color rgb="FFD1D5DB"/></right>
      <top style="thin"><color rgb="FFD1D5DB"/></top>
      <bottom style="thin"><color rgb="FFD1D5DB"/></bottom>
      <diagonal/>
    </border>
  </borders>
  <cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>
  <cellXfs count="6">
    <xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>
    <xf numFmtId="0" fontId="1" fillId="2" borderId="1" xfId="0" applyFont="1" applyFill="1" applyBorder="1" applyAlignment="1"><alignment horizontal="center" vertical="center"/></xf>
    <xf numFmtId="0" fontId="2" fillId="3" borderId="1" xfId="0" applyFont="1" applyFill="1" applyBorder="1" applyAlignment="1"><alignment horizontal="center" vertical="center" wrapText="1"/></xf>
    <xf numFmtId="0" fontId="3" fillId="4" borderId="1" xfId="0" applyFont="1" applyFill="1" applyBorder="1"/>
    <xf numFmtId="0" fontId="0" fillId="4" borderId="1" xfId="0" applyFill="1" applyBorder="1"/>
    <xf numFmtId="0" fontId="0" fillId="0" borderId="1" xfId="0" applyBorder="1" applyAlignment="1"><alignment vertical="top" wrapText="1"/></xf>
  </cellXfs>
  <cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>
</styleSheet>"""


def _legacy_xlsx_report_response(filename, title, columns, rows):
    data_rows = list(rows)
    generated_at = timezone.localtime().strftime("%d-%m-%Y %H:%M:%S")
    header_row = 5
    first_data_row = header_row + 1
    last_col = _xlsx_col_name(len(columns))
    last_row = max(header_row, first_data_row + len(data_rows) - 1)

    widths = []
    for index, column in enumerate(columns):
        max_length = len(str(column))
        for row in data_rows:
            if index < len(row):
                max_length = max(max_length, len("" if row[index] is None else str(row[index])))
        limit = 78 if "Detalle" in column else 42
        widths.append(min(max(max_length + 2, 12), limit))

    cols_xml = "".join(
        f'<col min="{index}" max="{index}" width="{width}" customWidth="1"/>'
        for index, width in enumerate(widths, start=1)
    )
    rows_xml = [
        _xlsx_row(1, [(title, 1)]),
        _xlsx_row(2, [("Generado", 3), (generated_at, 4)]),
        _xlsx_row(3, [("Total registros", 3), (len(data_rows), 4)]),
        _xlsx_row(4, [""] * len(columns)),
        _xlsx_row(header_row, [(column, 2) for column in columns]),
    ]

    for row_index, row in enumerate(data_rows, start=first_data_row):
        padded_row = list(row) + [""] * (len(columns) - len(row))
        rows_xml.append(_xlsx_row(row_index, padded_row[:len(columns)], 5))

    worksheet_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <dimension ref="A1:{last_col}{last_row}"/>
  <sheetViews>
    <sheetView workbookViewId="0">
      <pane ySplit="{header_row}" topLeftCell="A{first_data_row}" activePane="bottomLeft" state="frozen"/>
    </sheetView>
  </sheetViews>
  <cols>{cols_xml}</cols>
  <sheetData>{"".join(rows_xml)}</sheetData>
  <mergeCells count="1"><mergeCell ref="A1:{last_col}1"/></mergeCells>
  <autoFilter ref="A{header_row}:{last_col}{last_row}"/>
  <pageMargins left="0.7" right="0.7" top="0.75" bottom="0.75" header="0.3" footer="0.3"/>
</worksheet>"""

    workbook_xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets><sheet name="Reporte" sheetId="1" r:id="rId1"/></sheets>
</workbook>"""
    workbook_rels_xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>
</Relationships>"""
    rels_xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>
  <Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>
</Relationships>"""
    content_types_xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>
  <Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>
  <Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>
</Types>"""
    now_iso = timezone.now().isoformat()
    core_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" xmlns:dc="http://purl.org/dc/elements/1.1/" xmlns:dcterms="http://purl.org/dc/terms/" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
  <dc:title>{xml_escape(title)}</dc:title>
  <dc:creator>Panel de Sostenibilidad</dc:creator>
  <dcterms:created xsi:type="dcterms:W3CDTF">{now_iso}</dcterms:created>
  <dcterms:modified xsi:type="dcterms:W3CDTF">{now_iso}</dcterms:modified>
</cp:coreProperties>"""
    app_xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">
  <Application>Django</Application>
</Properties>"""

    output = BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", content_types_xml)
        archive.writestr("_rels/.rels", rels_xml)
        archive.writestr("docProps/core.xml", core_xml)
        archive.writestr("docProps/app.xml", app_xml)
        archive.writestr("xl/workbook.xml", workbook_xml)
        archive.writestr("xl/_rels/workbook.xml.rels", workbook_rels_xml)
        archive.writestr("xl/styles.xml", _xlsx_styles_xml())
        archive.writestr("xl/worksheets/sheet1.xml", worksheet_xml)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _xlsx_report_response(filename, title, columns, rows):
    """Genera un XLSX compacto, sin filas ni celdas de datos vacías."""
    data_rows = list(rows)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Reporte"
    header_row = 4
    first_data_row = header_row + 1
    last_column = len(columns)
    worksheet.freeze_panes = f"A{first_data_row}"

    title_fill = PatternFill("solid", fgColor="0891B2")
    header_fill = PatternFill("solid", fgColor="0F172A")
    meta_fill = PatternFill("solid", fgColor="EFF6FF")
    white_font = Font(color="FFFFFF", bold=True)
    meta_font = Font(color="334155", bold=True)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    title_cell = worksheet.cell(1, 1, title)
    title_cell.fill = title_fill
    title_cell.font = Font(color="FFFFFF", bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    generated_at = timezone.localtime().strftime("%d-%m-%Y %H:%M:%S")
    worksheet.cell(2, 1, "Generado")
    worksheet.cell(2, 2, generated_at)
    worksheet.cell(3, 1, "Total registros")
    worksheet.cell(3, 2, len(data_rows))
    if last_column > 2:
        worksheet.merge_cells(start_row=2, start_column=2, end_row=2, end_column=last_column)
        worksheet.merge_cells(start_row=3, start_column=2, end_row=3, end_column=last_column)
    for row_number in (2, 3):
        for col_number in range(1, last_column + 1):
            cell = worksheet.cell(row_number, col_number)
            cell.fill = meta_fill
            cell.border = border
        worksheet.cell(row_number, 1).font = meta_font

    for col_number, column in enumerate(columns, start=1):
        cell = worksheet.cell(header_row, col_number, str(column))
        cell.fill = header_fill
        cell.font = white_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    rows_to_write = data_rows or [["Sin registros para mostrar."]]
    for row_number, row in enumerate(rows_to_write, start=first_data_row):
        padded_row = list(row) + [None] * (last_column - len(row))
        for col_number, value in enumerate(padded_row[:last_column], start=1):
            if value is None or (isinstance(value, str) and not value.strip()):
                value = "Sin información"
            cell = worksheet.cell(row_number, col_number, value)
            # Fuerza los textos como texto para impedir que datos ingresados por
            # usuarios se interpreten como fórmulas de Excel.
            if isinstance(value, str):
                cell.data_type = "s"
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    last_row = header_row + len(rows_to_write)
    last_column_letter = get_column_letter(last_column)
    worksheet.auto_filter.ref = f"A{header_row}:{last_column_letter}{last_row}"
    if not data_rows and last_column > 1:
        worksheet.merge_cells(
            start_row=first_data_row,
            start_column=1,
            end_row=first_data_row,
            end_column=last_column,
        )
        empty_cell = worksheet.cell(first_data_row, 1)
        empty_cell.value = "Sin registros para mostrar."
        empty_cell.alignment = Alignment(horizontal="center", vertical="center")
        empty_cell.font = Font(color="64748B", italic=True)

    for col_number, column in enumerate(columns, start=1):
        values = [str(column)]
        for row in data_rows:
            if col_number - 1 < len(row):
                value = row[col_number - 1]
                values.append(
                    "Sin información"
                    if value is None or (isinstance(value, str) and not value.strip())
                    else str(value)
                )
        max_length = max(len(value) for value in values)
        limit = 78 if "Detalle" in str(column) else 42
        worksheet.column_dimensions[get_column_letter(col_number)].width = min(
            max(max_length + 2, 12),
            limit,
        )

    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response["Content-Length"] = len(output.getvalue())
    return response


def _residuos_reporte_columns():
    return [
        "ID", "Fecha registro", "Fecha confirmacion", "Categoria", "Subtipo",
        "Contenedor", "Peso kg", "Unidades", "Estado", "Retirado",
        "Asignatura", "Seccion", "Profesor", "Horario", "N clase",
        "Origen registro", "Motivo manual", "Creado en sistema",
        "Creado por", "Actualizado por", "IP origen",
    ]


def _retiros_reporte_columns():
    return [
        "ID", "Fecha retiro", "Usuario", "Origen", "Tipo residuo",
        "Destino", "Categoria destino", "Direccion destino",
        "Cantidad peso", "Cantidad unidades", "Detalle",
    ]


def _residuos_queryset_reporte(clase=None):
    queryset = (
        Residuo.objects
        .select_related("tipo", "subtipo", "created_by", "updated_by")
        .order_by("-hora_escaneo", "-id")
    )
    if clase is not None:
        filtros = {
            "asignatura": clase.asignatura,
            "seccion": clase.seccion,
            "profesor": clase.profesor,
        }
        inicio = _inicio_clase_datetime(clase)
        fin = _fin_clase_datetime(clase)
        if inicio and fin:
            filtros.update(hora_escaneo__gte=inicio, hora_escaneo__lte=fin)
        else:
            filtros["hora_escaneo__date"] = clase.fecha
        queryset = queryset.filter(**filtros)
    return queryset


def _retiros_queryset_reporte(clase=None):
    queryset = (
        HistorialRetiro.objects
        .select_related("usuario", "destino", "destino__categoria")
        .order_by("-fecha_retiro", "-id")
    )
    if clase is not None:
        inicio = _inicio_clase_datetime(clase)
        fin = _fin_clase_datetime(clase)
        if inicio and fin:
            queryset = queryset.filter(fecha_retiro__gte=inicio, fecha_retiro__lte=fin)
        else:
            queryset = queryset.filter(fecha_retiro__date=clase.fecha)
    return queryset


def _residuos_reporte_rows(residuos_qs):
    rows = []
    for residuo in residuos_qs:
        rows.append([
            residuo.id,
            _format_report_datetime(residuo.hora_escaneo),
            _format_report_datetime(residuo.confirmado_at),
            residuo.tipo.nombre if residuo.tipo else "",
            residuo.subtipo.nombre_residuo if residuo.subtipo else "",
            residuo.contenedor_id or "",
            residuo.peso if residuo.peso is not None else "",
            residuo.unidad if residuo.unidad is not None else "",
            residuo.get_estado_display(),
            "Si" if residuo.retirado else "No",
            residuo.asignatura,
            residuo.seccion,
            residuo.profesor,
            residuo.horario,
            residuo.numero_clase if residuo.numero_clase is not None else "",
            "Manual" if residuo.manual_entry else "Automatico",
            residuo.manual_reason or "",
            _format_report_datetime(residuo.created_at),
            residuo.created_by.username if residuo.created_by else "",
            residuo.updated_by.username if residuo.updated_by else "",
            residuo.source_ip or "",
        ])
    return rows


def _retiros_reporte_rows(retiros):
    rows = []
    for retiro in retiros:
        rows.append([
            retiro.id,
            _format_report_datetime(retiro.fecha_retiro),
            retiro.usuario.username if retiro.usuario else "",
            retiro.contenedor_origen,
            retiro.tipo_residuo,
            retiro.destino.nombre if retiro.destino else "",
            retiro.destino.categoria.nombre if retiro.destino and retiro.destino.categoria else "",
            retiro.destino.direccion if retiro.destino else "",
            retiro.cantidad_peso,
            retiro.cantidad_unidades,
            retiro.detalle or "",
        ])
    return rows


def _crear_reporte_residuos(request):
    rows = _residuos_reporte_rows(_residuos_queryset_reporte())
    _audit(request, "residuos_excel_exported", metadata={"total": len(rows)})
    return _xlsx_report_response("reporte_residuos.xlsx", "Reporte de Residuos", _residuos_reporte_columns(), rows)


def _crear_reporte_retiros(request):
    rows = _retiros_reporte_rows(_retiros_queryset_reporte())

    _audit(request, "retiros_excel_exported", metadata={"total": len(rows)})
    return _xlsx_report_response("reporte_retiros.xlsx", "Reporte de Retiros", _retiros_reporte_columns(), rows)


def _xlsx_report_bytes(filename, title, columns, rows):
    return _xlsx_report_response(filename, title, columns, rows).content


def _exportacion_taller_paths(clase, filename):
    nombre_taller = slugify(f"{clase.asignatura}-{clase.seccion}")[:80] or "taller"
    relative_dir = Path(
        f"{clase.fecha:%Y-%m-%d}_{nombre_taller}_taller-{clase.pk}"
    )
    relative_path = relative_dir / filename
    absolute_dir = settings.EXPORT_DIR / relative_dir
    absolute_dir.mkdir(parents=True, exist_ok=True)
    return relative_path, absolute_dir / filename


def _guardar_reporte_taller(clase, tipo, force=False):
    if tipo == "residuos":
        filename = "exportar residuos.xlsx"
        title = f"Exportar residuos - {clase.asignatura}"
        columns = _residuos_reporte_columns()
        rows = _residuos_reporte_rows(_residuos_queryset_reporte(clase))
    elif tipo == "retiros":
        filename = "exportar retiros.xlsx"
        title = f"Exportar retiros - {clase.asignatura}"
        columns = _retiros_reporte_columns()
        rows = _retiros_reporte_rows(_retiros_queryset_reporte(clase))
    else:
        raise ValueError("Tipo de exportación no válido.")

    relative_path, absolute_path = _exportacion_taller_paths(clase, filename)
    if force or not absolute_path.exists():
        absolute_path.write_bytes(_xlsx_report_bytes(filename, title, columns, rows))
    return relative_path.as_posix(), len(rows)


def _export_path_exists(relative_path):
    if not relative_path:
        return False
    return (settings.EXPORT_DIR / relative_path).exists()


def _generar_exportaciones_taller(clase, request=None, force=False):
    clase.refresh_from_db()
    if (
        not force
        and clase.exportado
        and _export_path_exists(clase.export_residuos_path)
        and _export_path_exists(clase.export_retiros_path)
    ):
        return False

    residuos_path, total_residuos = _guardar_reporte_taller(clase, "residuos", force=force)
    retiros_path, total_retiros = _guardar_reporte_taller(clase, "retiros", force=force)

    clase.export_residuos_path = residuos_path
    clase.export_retiros_path = retiros_path
    clase.exportado = True
    clase.exportado_at = timezone.now()
    clase.export_error = ""
    clase.export_residuos_descargado_at = None
    clase.export_retiros_descargado_at = None
    clase.save(update_fields=[
        "export_residuos_path",
        "export_retiros_path",
        "exportado",
        "exportado_at",
        "export_error",
        "export_residuos_descargado_at",
        "export_retiros_descargado_at",
    ])

    _audit(request, "taller_exports_generated", clase, {
        "residuos": total_residuos,
        "retiros": total_retiros,
        "force": force,
        "carpeta": str((settings.EXPORT_DIR / Path(residuos_path).parent).resolve()),
    })
    return True


def _registrar_error_exportacion(clase, request, error):
    mensaje = str(error)[:2000] or error.__class__.__name__
    ClaseHorario.objects.filter(pk=clase.pk).update(
        exportado=False,
        export_error=mensaje,
    )
    logger.exception("Error generando exportaciones del taller %s: %s", clase.pk, error)
    _audit(request, "taller_exports_failed", clase, {"error": mensaje})


def _clase_exportacion_queryset(user):
    queryset = ClaseHorario.objects.exclude(profesor=PROFESOR_TALLERES_PRECARGADOS)
    if not es_administrador(user):
        queryset = queryset.filter(profesor=user.username)
    return queryset


@login_required(login_url="login")
@require_POST
def reintentar_exportacion_taller(request, id):
    if not puede_gestionar_horarios(request.user):
        messages.error(request, "No tiene permisos para reintentar exportaciones.")
        return redirect("dashboard")

    clase = get_object_or_404(_clase_exportacion_queryset(request.user), pk=id)
    try:
        _generar_exportaciones_taller(clase, request=request, force=True)
    except Exception as exc:
        _registrar_error_exportacion(clase, request, exc)
        messages.error(request, "No se pudo generar la exportación. El error quedó registrado.")
    else:
        messages.success(request, "Exportaciones generadas correctamente.")

    return redirect(_safe_next_url(request) or "profesor")


@admin_required
@require_GET
def exportar_residuos_excel(request):
    return _crear_reporte_residuos(request)


@admin_required
@require_GET
def exportar_retiros_excel(request):
    return _crear_reporte_retiros(request)


@admin_required
@require_POST
def generar_backup_db(request):
    engine = settings.DATABASES["default"]["ENGINE"]

    if engine != "django.db.backends.sqlite3":
        messages.error(request, "El backup automático está disponible solo para SQLite.")
        return redirect("dashboard")

    try:
        destino = create_rolling_backup()
    except (OSError, ValueError) as exc:
        logger.exception("No se pudo generar el backup verificado: %s", exc)
        messages.error(request, "No se pudo crear un respaldo íntegro y verificable.")
        return redirect("dashboard")

    try:
        try:
            audit_path = str(destino.relative_to(settings.BASE_DIR))
        except ValueError:
            audit_path = str(destino)
    except ValueError:
        audit_path = str(destino)

    _audit(request, "database_rolling_backup_updated", metadata={
        "path": audit_path,
        "verified": True,
        "modo": "unico_actualizable",
        "motivo": "solicitud_manual_administrador",
    })
    messages.success(request, f"Respaldo único actualizado: {destino.name}")
    return redirect("dashboard")


@require_GET
def health_check(request):
    user = getattr(request, "user", None)
    if settings.HEALTH_CHECK_REQUIRE_LOGIN and not getattr(user, "is_authenticated", False):
        return JsonResponse({"ok": False, "error": "Autenticacion requerida."}, status=403)

    payload = {"ok": True}
    if settings.HEALTH_CHECK_EXPOSE_DETAILS or (user is not None and es_administrador(user)):
        payload.update({
            "debug": settings.DEBUG,
            "database": settings.DATABASES["default"]["ENGINE"],
        })
    return JsonResponse(payload)
